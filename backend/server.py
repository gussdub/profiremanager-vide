from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import json
import hashlib
import re
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from io import BytesIO
import base64
from sendgrid import SendGridAPIClient
from sendgrid.helpers.mail import Mail

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI(title="ProFireManager API", version="2.0")

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# JWT and Password configuration
SECRET_KEY = os.environ.get("JWT_SECRET", "your-secret-key-here")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 24 * 60  # 24 hours

# Simplified password hashing

# Helper functions
def validate_complex_password(password: str) -> bool:
    """
    Valide qu'un mot de passe respecte les critères de complexité :
    - 8 caractères minimum
    - 1 majuscule
    - 1 chiffre  
    - 1 caractère spécial (!@#$%^&*+-?())
    """
    if len(password) < 8:
        return False
    
    has_uppercase = bool(re.search(r'[A-Z]', password))
    has_digit = bool(re.search(r'\d', password))
    has_special = bool(re.search(r'[!@#$%^&*+\-?()]', password))
    
    return has_uppercase and has_digit and has_special

def send_welcome_email(user_email: str, user_name: str, user_role: str, temp_password: str):
    """
    Envoie un email de bienvenue avec les informations de connexion
    """
    try:
        # Définir les modules selon le rôle
        modules_by_role = {
            'admin': [
                "📊 Tableau de bord - Vue d'ensemble et statistiques",
                "👥 Personnel - Gestion complète des pompiers", 
                "📅 Planning - Attribution automatique et manuelle",
                "🔄 Remplacements - Validation des demandes",
                "📚 Formations - Inscription et gestion",
                "📈 Rapports - Analyses et exports",
                "⚙️ Paramètres - Configuration système",
                "👤 Mon profil - Informations personnelles"
            ],
            'superviseur': [
                "📊 Tableau de bord - Vue d'ensemble et statistiques",
                "👥 Personnel - Consultation des pompiers",
                "📅 Planning - Gestion et validation", 
                "🔄 Remplacements - Approbation des demandes",
                "📚 Formations - Inscription et gestion",
                "👤 Mon profil - Informations personnelles"
            ],
            'employe': [
                "📊 Tableau de bord - Vue d'ensemble personnalisée",
                "📅 Planning - Consultation de votre planning",
                "🔄 Remplacements - Demandes de remplacement",
                "📚 Formations - Inscription aux formations",
                "👤 Mon profil - Informations et disponibilités"
            ]
        }
        
        role_name = {
            'admin': 'Administrateur',
            'superviseur': 'Superviseur', 
            'employe': 'Employé'
        }.get(user_role, 'Utilisateur')
        
        user_modules = modules_by_role.get(user_role, modules_by_role['employe'])
        modules_html = ''.join([f'<li style="margin-bottom: 8px;">{module}</li>' for module in user_modules])
        
        subject = f"Bienvenue dans ProFireManager v2.0 - Votre compte {role_name}"
        
        html_content = f"""
        <html>
        <body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
            <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
                <div style="text-align: center; margin-bottom: 30px;">
                    <img src="https://customer-assets.emergentagent.com/job_fireshift-manager/artifacts/6vh2i9cz_05_Icone_Flamme_Rouge_Bordure_D9072B_VISIBLE.png" 
                         alt="ProFireManager" 
                         style="width: 100px; height: 100px; margin-bottom: 15px;">
                    <h1 style="color: #dc2626; margin: 0;">ProFireManager v2.0</h1>
                    <p style="color: #666; margin: 5px 0;">Système de gestion des services d'incendie</p>
                </div>
                
                <h2 style="color: #1e293b;">Bonjour {user_name},</h2>
                
                <p>Votre compte <strong>{role_name}</strong> a été créé avec succès dans ProFireManager v2.0, le système de gestion des horaires et remplacements automatisés pour les services d'incendie du Canada.</p>
                
                <div style="background: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; padding: 20px; margin: 20px 0;">
                    <h3 style="color: #dc2626; margin-top: 0;">🔑 Informations de connexion :</h3>
                    <p><strong>Email :</strong> {user_email}</p>
                    <p><strong>Mot de passe temporaire :</strong> {temp_password}</p>
                    <p style="color: #dc2626; font-weight: bold;">⚠️ Veuillez modifier votre mot de passe lors de votre première connexion</p>
                </div>
                
                <div style="text-align: center; margin: 30px 0;">
                    <a href="{os.environ.get('FRONTEND_URL', 'https://epimanager-1.preview.emergentagent.com')}" 
                       style="background: #dc2626; color: white; padding: 12px 24px; text-decoration: none; border-radius: 8px; font-weight: bold; display: inline-block;">
                        🚒 Accéder à ProFireManager
                    </a>
                    <p style="font-size: 12px; color: #666; margin-top: 10px;">
                        💡 Conseil : Ajoutez ce lien à vos favoris pour un accès rapide
                    </p>
                </div>
                
                <h3 style="color: #1e293b;">📋 Modules disponibles pour votre rôle ({role_name}) :</h3>
                <ul style="background: #f0fdf4; border-left: 4px solid #10b981; padding: 15px 20px; margin: 15px 0;">
                    {modules_html}
                </ul>
                
                <div style="background: #fef3c7; border: 1px solid #fcd34d; border-radius: 8px; padding: 15px; margin: 20px 0;">
                    <h4 style="color: #92400e; margin-top: 0;">🔒 Sécurité de votre compte :</h4>
                    <ul style="color: #78350f; margin: 10px 0;">
                        <li>Modifiez votre mot de passe temporaire dès votre première connexion</li>
                        <li>Utilisez un mot de passe complexe (8 caractères, majuscule, chiffre, caractère spécial)</li>
                        <li>Ne partagez jamais vos identifiants</li>
                        <li>Déconnectez-vous après chaque session</li>
                    </ul>
                </div>
                
                <hr style="border: none; border-top: 1px solid #e2e8f0; margin: 30px 0;">
                
                <p style="color: #666; font-size: 14px; text-align: center;">
                    Cet email a été envoyé automatiquement par ProFireManager v2.0.<br>
                    Si vous avez des questions, contactez votre administrateur système.
                </p>
                
                <div style="text-align: center; margin-top: 20px;">
                    <p style="color: #999; font-size: 12px;">
                        ProFireManager v2.0 - Système de gestion des services d'incendie du Canada<br>
                        Développé pour optimiser la gestion des horaires et remplacements automatisés
                    </p>
                </div>
            </div>
        </body>
        </html>
        """
        
        # Envoyer l'email via SendGrid
        sendgrid_api_key = os.environ.get('SENDGRID_API_KEY')
        sender_email = os.environ.get('SENDER_EMAIL', 'noreply@profiremanager.ca')
        
        if not sendgrid_api_key:
            print(f"[WARNING] SENDGRID_API_KEY non configurée - Email non envoyé à {user_email}")
            return False
        
        try:
            message = Mail(
                from_email=sender_email,
                to_emails=user_email,
                subject=subject,
                html_content=html_content
            )
            
            sg = SendGridAPIClient(sendgrid_api_key)
            response = sg.send(message)
            
            if response.status_code in [200, 201, 202]:
                print(f"✅ Email de bienvenue envoyé avec succès à {user_email}")
                return True
            else:
                print(f"⚠️ Erreur SendGrid (code {response.status_code}) pour {user_email}")
                return False
                
        except Exception as e:
            print(f"❌ Erreur lors de l'envoi de l'email à {user_email}: {str(e)}")
            return False
        
    except Exception as e:
        print(f"Erreur envoi email: {str(e)}")
        return False

def verify_password(plain_password, hashed_password):
    return hashlib.sha256(plain_password.encode()).hexdigest() == hashed_password

def get_password_hash(password):
    return hashlib.sha256(password.encode()).hexdigest()

security = HTTPBearer()

# Define Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    prenom: str
    email: str
    telephone: str = ""
    contact_urgence: str = ""
    grade: str  # Capitaine, Directeur, Pompier, Lieutenant
    fonction_superieur: bool = False  # Pour pompiers pouvant agir comme lieutenant
    type_emploi: str  # temps_plein, temps_partiel
    heures_max_semaine: int = 40  # Heures max par semaine (pour temps partiel)
    role: str  # admin, superviseur, employe
    statut: str = "Actif"  # Actif, Inactif
    numero_employe: str
    date_embauche: str
    formations: List[str] = []
    mot_de_passe_hash: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    nom: str
    prenom: str
    email: str
    telephone: str = ""
    contact_urgence: str = ""
    grade: str
    fonction_superieur: bool = False
    type_emploi: str
    heures_max_semaine: int = 40
    role: str
    numero_employe: str
    date_embauche: str
    formations: List[str] = []
    mot_de_passe: str

class UserLogin(BaseModel):
    email: str
    mot_de_passe: str

class TypeGarde(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    heure_debut: str
    heure_fin: str
    personnel_requis: int
    duree_heures: int
    couleur: str
    jours_application: List[str] = []  # monday, tuesday, etc.
    officier_obligatoire: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TypeGardeCreate(BaseModel):
    nom: str
    heure_debut: str
    heure_fin: str
    personnel_requis: int
    duree_heures: int
    couleur: str
    jours_application: List[str] = []
    officier_obligatoire: bool = False

class Planning(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    semaine_debut: str  # Format: YYYY-MM-DD
    semaine_fin: str
    assignations: Dict[str, Any] = {}  # jour -> type_garde -> assignation
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PlanningCreate(BaseModel):
    semaine_debut: str
    semaine_fin: str
    assignations: Dict[str, Any] = {}

class Assignation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    type_garde_id: str
    date: str
    statut: str = "planifie"  # planifie, confirme, remplacement_demande
    assignation_type: str = "auto"  # auto, manuel, manuel_avance

class AssignationCreate(BaseModel):
    user_id: str
    type_garde_id: str
    date: str
    assignation_type: str = "manuel"

class DemandeRemplacement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    demandeur_id: str
    type_garde_id: str
    date: str
    raison: str
    statut: str = "en_cours"  # en_cours, approuve, refuse
    remplacant_id: Optional[str] = None
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DemandeRemplacementCreate(BaseModel):
    type_garde_id: str
    date: str
    raison: str

class Formation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    nom: str
    description: str = ""
    duree_heures: int = 0
    validite_mois: int = 12  # 0 = Pas de renouvellement
    obligatoire: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class FormationCreate(BaseModel):
    nom: str
    description: str = ""
    duree_heures: int = 0
    validite_mois: int = 12  # 0 = Pas de renouvellement
    obligatoire: bool = False

class Disponibilite(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    date: str  # Date exacte YYYY-MM-DD
    type_garde_id: Optional[str] = None  # Spécifier pour quel type de garde
    heure_debut: str
    heure_fin: str
    statut: str = "disponible"  # disponible, indisponible, preference
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DisponibiliteCreate(BaseModel):
    user_id: str
    date: str  # Date exacte YYYY-MM-DD
    type_garde_id: Optional[str] = None
    heure_debut: str
    heure_fin: str
    statut: str = "disponible"

class Statistiques(BaseModel):
    personnel_actif: int
    gardes_cette_semaine: int
    formations_planifiees: int
    taux_couverture: float
    heures_travaillees: int
    remplacements_effectues: int

# Helper functions

def create_access_token(data: dict, expires_delta: Optional[timedelta] = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        payload = jwt.decode(credentials.credentials, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Token invalide")
    except jwt.PyJWTError:
        raise HTTPException(status_code=401, detail="Token invalide")
    
    user = await db.users.find_one({"id": user_id})
    if user is None:
        raise HTTPException(status_code=401, detail="Utilisateur non trouvé")
    return User(**user)

# Root route
@api_router.get("/")
async def root():
    return {"message": "ProFireManager API v2.0", "status": "running"}
@api_router.post("/auth/login")
async def login(user_login: UserLogin):
    user_data = await db.users.find_one({"email": user_login.email})
    if not user_data or not verify_password(user_login.mot_de_passe, user_data["mot_de_passe_hash"]):
        raise HTTPException(status_code=401, detail="Email ou mot de passe incorrect")
    
    user = User(**user_data)
    access_token = create_access_token(data={"sub": user.id})
    return {
        "access_token": access_token,
        "token_type": "bearer",
        "user": {
            "id": user.id,
            "nom": user.nom,
            "prenom": user.prenom,
            "email": user.email,
            "role": user.role,
            "grade": user.grade,
            "type_emploi": user.type_emploi
        }
    }

@api_router.get("/auth/me")
async def get_current_user_info(current_user: User = Depends(get_current_user)):
    return {
        "id": current_user.id,
        "nom": current_user.nom,
        "prenom": current_user.prenom,
        "email": current_user.email,
        "role": current_user.role,
        "grade": current_user.grade,
        "type_emploi": current_user.type_emploi,
        "formations": current_user.formations
    }

# User management routes
@api_router.post("/users", response_model=User)
async def create_user(user_create: UserCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Validation du mot de passe complexe
    if not validate_complex_password(user_create.mot_de_passe):
        raise HTTPException(
            status_code=400, 
            detail="Le mot de passe doit contenir au moins 8 caractères, une majuscule, un chiffre et un caractère spécial (!@#$%^&*+-?())"
        )
    
    # Check if user already exists
    existing_user = await db.users.find_one({"email": user_create.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Cet email est déjà utilisé")
    
    user_dict = user_create.dict()
    temp_password = user_dict["mot_de_passe"]  # Sauvegarder pour l'email
    user_dict["mot_de_passe_hash"] = get_password_hash(user_dict.pop("mot_de_passe"))
    user_obj = User(**user_dict)
    
    await db.users.insert_one(user_obj.dict())
    
    # Envoyer l'email de bienvenue
    try:
        user_name = f"{user_create.prenom} {user_create.nom}"
        email_sent = send_welcome_email(user_create.email, user_name, user_create.role, temp_password)
        
        if email_sent:
            print(f"Email de bienvenue envoyé à {user_create.email}")
        else:
            print(f"Échec envoi email à {user_create.email}")
            
    except Exception as e:
        print(f"Erreur lors de l'envoi de l'email: {str(e)}")
        # Ne pas échouer la création du compte si l'email échoue
    
    return user_obj

@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    users = await db.users.find().to_list(1000)
    cleaned_users = [clean_mongo_doc(user) for user in users]
    return [User(**user) for user in cleaned_users]

@api_router.get("/users/{user_id}", response_model=User)
async def get_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    user = await db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    user = clean_mongo_doc(user)
    return User(**user)

class ProfileUpdate(BaseModel):
    prenom: str
    nom: str
    email: str
    telephone: str = ""
    contact_urgence: str = ""
    heures_max_semaine: int = 25

@api_router.put("/users/mon-profil", response_model=User)
async def update_mon_profil(
    profile_data: ProfileUpdate,
    current_user: User = Depends(get_current_user)
):
    """
    Permet à un utilisateur de modifier son propre profil
    """
    try:
        # L'utilisateur peut modifier son propre profil
        result = await db.users.update_one(
            {"id": current_user.id}, 
            {"$set": profile_data.dict()}
        )
        
        if result.modified_count == 0:
            raise HTTPException(status_code=400, detail="Impossible de mettre à jour le profil")
        
        # Récupérer le profil mis à jour
        updated_user = await db.users.find_one({"id": current_user.id})
        updated_user = clean_mongo_doc(updated_user)
        return User(**updated_user)
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur mise à jour profil: {str(e)}")

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_update: UserCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Update user data
    user_dict = user_update.dict()
    if user_dict["mot_de_passe"]:
        user_dict["mot_de_passe_hash"] = get_password_hash(user_dict.pop("mot_de_passe"))
    else:
        user_dict.pop("mot_de_passe")
        user_dict["mot_de_passe_hash"] = existing_user["mot_de_passe_hash"]
    
    user_dict["id"] = user_id
    user_dict["statut"] = existing_user.get("statut", "Actif")
    user_dict["created_at"] = existing_user.get("created_at")
    
    result = await db.users.replace_one({"id": user_id}, user_dict)
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de mettre à jour l'utilisateur")
    
    updated_user = await db.users.find_one({"id": user_id})
    updated_user = clean_mongo_doc(updated_user)
    return User(**updated_user)

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Delete user
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de supprimer l'utilisateur")
    
    # Also delete related data
    await db.disponibilites.delete_many({"user_id": user_id})
    await db.assignations.delete_many({"user_id": user_id})
    await db.demandes_remplacement.delete_many({"demandeur_id": user_id})
    
    return {"message": "Utilisateur supprimé avec succès"}

@api_router.put("/users/{user_id}/access", response_model=User)
async def update_user_access(user_id: str, role: str, statut: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Validation des valeurs
    valid_roles = ["admin", "superviseur", "employe"]
    valid_statuts = ["Actif", "Inactif"]
    
    if role not in valid_roles:
        raise HTTPException(status_code=400, detail="Rôle invalide")
    if statut not in valid_statuts:
        raise HTTPException(status_code=400, detail="Statut invalide")
    
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Update user access
    result = await db.users.update_one(
        {"id": user_id}, 
        {"$set": {"role": role, "statut": statut}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de mettre à jour l'accès")
    
    updated_user = await db.users.find_one({"id": user_id})
    updated_user = clean_mongo_doc(updated_user)
    return User(**updated_user)

@api_router.delete("/users/{user_id}/revoke")
async def revoke_user_completely(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="Utilisateur non trouvé")
    
    # Prevent admin from deleting themselves
    if user_id == current_user.id:
        raise HTTPException(status_code=400, detail="Impossible de supprimer votre propre compte")
    
    # Delete user and all related data
    await db.users.delete_one({"id": user_id})
    await db.disponibilites.delete_many({"user_id": user_id})
    await db.assignations.delete_many({"user_id": user_id})
    await db.demandes_remplacement.delete_many({"demandeur_id": user_id})
    await db.demandes_remplacement.delete_many({"remplacant_id": user_id})
    
    return {"message": "Utilisateur et toutes ses données ont été supprimés définitivement"}

# Types de garde routes
@api_router.post("/types-garde", response_model=TypeGarde)
async def create_type_garde(type_garde: TypeGardeCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    type_garde_obj = TypeGarde(**type_garde.dict())
    await db.types_garde.insert_one(type_garde_obj.dict())
    return type_garde_obj

@api_router.get("/types-garde", response_model=List[TypeGarde])
async def get_types_garde(current_user: User = Depends(get_current_user)):
    types_garde = await db.types_garde.find().to_list(1000)
    cleaned_types = [clean_mongo_doc(type_garde) for type_garde in types_garde]
    return [TypeGarde(**type_garde) for type_garde in cleaned_types]

# Helper function to clean MongoDB documents
def clean_mongo_doc(doc):
    """Remove MongoDB ObjectId and other non-serializable fields"""
    if doc and "_id" in doc:
        doc.pop("_id", None)
    return doc

@api_router.put("/types-garde/{type_garde_id}", response_model=TypeGarde)
async def update_type_garde(type_garde_id: str, type_garde_update: TypeGardeCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if type garde exists
    existing_type = await db.types_garde.find_one({"id": type_garde_id})
    if not existing_type:
        raise HTTPException(status_code=404, detail="Type de garde non trouvé")
    
    # Update type garde data
    type_dict = type_garde_update.dict()
    type_dict["id"] = type_garde_id
    type_dict["created_at"] = existing_type.get("created_at")
    
    result = await db.types_garde.replace_one({"id": type_garde_id}, type_dict)
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de mettre à jour le type de garde")
    
    updated_type = await db.types_garde.find_one({"id": type_garde_id})
    updated_type = clean_mongo_doc(updated_type)
    return TypeGarde(**updated_type)

@api_router.delete("/types-garde/{type_garde_id}")
async def delete_type_garde(type_garde_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if type garde exists
    existing_type = await db.types_garde.find_one({"id": type_garde_id})
    if not existing_type:
        raise HTTPException(status_code=404, detail="Type de garde non trouvé")
    
    # Delete type garde
    result = await db.types_garde.delete_one({"id": type_garde_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de supprimer le type de garde")
    
    # Also delete related assignations
    await db.assignations.delete_many({"type_garde_id": type_garde_id})
    
    return {"message": "Type de garde supprimé avec succès"}
@api_router.get("/planning/{semaine_debut}")
async def get_planning(semaine_debut: str, current_user: User = Depends(get_current_user)):
    planning = await db.planning.find_one({"semaine_debut": semaine_debut})
    if not planning:
        # Create empty planning for the week
        semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        planning_obj = Planning(semaine_debut=semaine_debut, semaine_fin=semaine_fin)
        await db.planning.insert_one(planning_obj.dict())
        planning = planning_obj.dict()
    else:
        planning = clean_mongo_doc(planning)
    
    return planning

@api_router.delete("/planning/assignation/{assignation_id}")
async def retirer_assignation(assignation_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Trouver l'assignation
        assignation = await db.assignations.find_one({"id": assignation_id})
        if not assignation:
            raise HTTPException(status_code=404, detail="Assignation non trouvée")
        
        # Supprimer l'assignation
        result = await db.assignations.delete_one({"id": assignation_id})
        if result.deleted_count == 0:
            raise HTTPException(status_code=400, detail="Impossible de retirer l'assignation")
        
        return {
            "message": "Assignation retirée avec succès",
            "assignation_supprimee": assignation_id,
            "date": assignation["date"],
            "user_id": assignation["user_id"]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur suppression assignation: {str(e)}")

@api_router.post("/planning/assignation")
async def create_assignation(assignation: AssignationCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Store assignation in database
    assignation_obj = Assignation(**assignation.dict())
    await db.assignations.insert_one(assignation_obj.dict())
    
    # Créer notification pour l'employé assigné
    user_assigne = await db.users.find_one({"id": assignation.user_id})
    type_garde = await db.types_garde.find_one({"id": assignation.type_garde_id})
    
    if user_assigne and type_garde:
        await creer_notification(
            destinataire_id=assignation.user_id,
            type="planning_assigne",
            titre="Nouveau quart assigné",
            message=f"Vous avez été assigné(e) au quart '{type_garde['nom']}' le {assignation.date}",
            lien="/planning",
            data={
                "assignation_id": assignation_obj.id,
                "date": assignation.date,
                "type_garde": type_garde["nom"]
            }
        )
    
    return {"message": "Assignation créée avec succès"}

@api_router.get("/planning/assignations/{semaine_debut}")
async def get_assignations(semaine_debut: str, current_user: User = Depends(get_current_user)):
    semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
    
    assignations = await db.assignations.find({
        "date": {
            "$gte": semaine_debut,
            "$lte": semaine_fin
        }
    }).to_list(1000)
    
    # Clean MongoDB documents
    cleaned_assignations = [clean_mongo_doc(assignation) for assignation in assignations]
    return [Assignation(**assignation) for assignation in cleaned_assignations]

# Remplacements routes
@api_router.post("/remplacements", response_model=DemandeRemplacement)
async def create_demande_remplacement(demande: DemandeRemplacementCreate, current_user: User = Depends(get_current_user)):
    demande_obj = DemandeRemplacement(**demande.dict(), demandeur_id=current_user.id)
    await db.demandes_remplacement.insert_one(demande_obj.dict())
    
    # Créer notification pour les superviseurs/admins
    superviseurs_admins = await db.users.find({"role": {"$in": ["superviseur", "admin"]}}).to_list(100)
    for user in superviseurs_admins:
        await creer_notification(
            destinataire_id=user["id"],
            type="remplacement_demande",
            titre="Nouvelle demande de remplacement",
            message=f"{current_user.prenom} {current_user.nom} demande un remplacement le {demande.date}",
            lien="/remplacements",
            data={"demande_id": demande_obj.id}
        )
    
    # Clean the object before returning to avoid ObjectId serialization issues
    cleaned_demande = clean_mongo_doc(demande_obj.dict())
    return DemandeRemplacement(**cleaned_demande)

@api_router.get("/remplacements", response_model=List[DemandeRemplacement])
async def get_demandes_remplacement(current_user: User = Depends(get_current_user)):
    if current_user.role == "employe":
        demandes = await db.demandes_remplacement.find({"demandeur_id": current_user.id}).to_list(1000)
    else:
        demandes = await db.demandes_remplacement.find().to_list(1000)
    
    cleaned_demandes = [clean_mongo_doc(demande) for demande in demandes]
    return [DemandeRemplacement(**demande) for demande in cleaned_demandes]

# Formations routes
@api_router.post("/formations", response_model=Formation)
async def create_formation(formation: FormationCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    formation_obj = Formation(**formation.dict())
    await db.formations.insert_one(formation_obj.dict())
    return formation_obj

@api_router.get("/formations", response_model=List[Formation])
async def get_formations(current_user: User = Depends(get_current_user)):
    formations = await db.formations.find().to_list(1000)
    cleaned_formations = [clean_mongo_doc(formation) for formation in formations]
    return [Formation(**formation) for formation in cleaned_formations]

@api_router.put("/formations/{formation_id}", response_model=Formation)
async def update_formation(formation_id: str, formation_update: FormationCreate, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if formation exists
    existing_formation = await db.formations.find_one({"id": formation_id})
    if not existing_formation:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    
    # Update formation data
    formation_dict = formation_update.dict()
    formation_dict["id"] = formation_id
    formation_dict["created_at"] = existing_formation.get("created_at")
    
    result = await db.formations.replace_one({"id": formation_id}, formation_dict)
    if result.modified_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de mettre à jour la formation")
    
    updated_formation = await db.formations.find_one({"id": formation_id})
    updated_formation = clean_mongo_doc(updated_formation)
    return Formation(**updated_formation)

@api_router.delete("/formations/{formation_id}")
async def delete_formation(formation_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Check if formation exists
    existing_formation = await db.formations.find_one({"id": formation_id})
    if not existing_formation:
        raise HTTPException(status_code=404, detail="Formation non trouvée")
    
    # Delete formation
    result = await db.formations.delete_one({"id": formation_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de supprimer la formation")
    
    # Remove from users' formations arrays
    await db.users.update_many(
        {"formations": formation_id},
        {"$pull": {"formations": formation_id}}
    )
    
    return {"message": "Formation supprimée avec succès"}

class SessionFormation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    titre: str
    competence_id: str  # Lien vers Formation (compétence)
    duree_heures: int
    date_debut: str  # YYYY-MM-DD
    heure_debut: str  # HH:MM
    lieu: str
    formateur: str
    descriptif: str
    plan_cours: str = ""
    places_max: int = 20
    participants: List[str] = []  # IDs des utilisateurs inscrits
    statut: str = "planifie"  # planifie, en_cours, termine, annule
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class SessionFormationCreate(BaseModel):
    titre: str
    competence_id: str
    duree_heures: int
    date_debut: str
    heure_debut: str
    lieu: str
    formateur: str
    descriptif: str
    plan_cours: str = ""
    places_max: int = 20

class InscriptionFormation(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    session_id: str
    user_id: str
    date_inscription: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    statut: str = "inscrit"  # inscrit, present, absent, annule

class DemandeCongé(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    demandeur_id: str
    type_conge: str  # maladie, vacances, parental, personnel
    date_debut: str  # YYYY-MM-DD
    date_fin: str  # YYYY-MM-DD
    nombre_jours: int
    raison: str
    documents: List[str] = []  # URLs des documents justificatifs
    priorite: str = "normale"  # urgente, haute, normale, faible
    statut: str = "en_attente"  # en_attente, approuve, refuse
    approuve_par: Optional[str] = None  # ID du superviseur/admin qui approuve
    date_approbation: Optional[str] = None
    commentaire_approbation: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class DemandeCongeCreate(BaseModel):
    type_conge: str
    date_debut: str
    date_fin: str
    raison: str
    priorite: str = "normale"
    documents: List[str] = []

class Notification(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    destinataire_id: str
    type: str  # remplacement_disponible, conge_approuve, conge_refuse, conge_demande, planning_assigne
    titre: str
    message: str
    lien: Optional[str] = None  # Lien vers la page concernée
    statut: str = "non_lu"  # non_lu, lu
    data: Optional[Dict[str, Any]] = {}  # Données supplémentaires (demande_id, etc.)
    date_creation: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    date_lecture: Optional[str] = None

class NotificationRemplacement(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    demande_remplacement_id: str
    destinataire_id: str
    message: str
    type_notification: str = "remplacement_disponible"  # remplacement_disponible, approbation_requise
    statut: str = "envoye"  # envoye, lu, accepte, refuse
    date_envoi: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    date_reponse: Optional[datetime] = None
    ordre_priorite: Optional[int] = None  # Pour le mode séquentiel

class ParametresRemplacements(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    mode_notification: str = "simultane"  # simultane, sequentiel, groupe_sequentiel
    taille_groupe: int = 3  # Pour mode groupe_sequentiel
    delai_attente_heures: int = 24  # Délai avant de passer au suivant
    max_contacts: int = 5
    priorite_grade: bool = True
    priorite_competences: bool = True

# EPI Models
class EPIEmploye(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    employe_id: str
    type_epi: str  # casque, bottes, veste_bunker, pantalon_bunker, gants, masque_scba
    taille: str
    date_attribution: str
    etat: str = "Neuf"  # Neuf, Bon, À remplacer, Défectueux
    date_expiration: str
    date_prochaine_inspection: Optional[str] = None
    historique_inspections: List[Dict[str, Any]] = []
    notes: Optional[str] = None
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())

class EPIEmployeCreate(BaseModel):
    employe_id: str
    type_epi: str
    taille: str
    date_attribution: str
    etat: str = "Neuf"
    date_expiration: str
    date_prochaine_inspection: Optional[str] = None
    notes: Optional[str] = None

class EPIEmployeUpdate(BaseModel):
    taille: Optional[str] = None
    etat: Optional[str] = None
    date_expiration: Optional[str] = None
    date_prochaine_inspection: Optional[str] = None
    notes: Optional[str] = None

class InspectionEPI(BaseModel):
    date_inspection: str
    inspecteur_id: str
    resultat: str  # Conforme, Non conforme, Remplacement nécessaire
    observations: Optional[str] = None
    photos: List[str] = []  # URLs des photos d'inspection

# Demandes de congé routes
@api_router.post("/demandes-conge", response_model=DemandeCongé)
async def create_demande_conge(demande: DemandeCongeCreate, current_user: User = Depends(get_current_user)):
    # Calculer le nombre de jours
    date_debut = datetime.strptime(demande.date_debut, "%Y-%m-%d")
    date_fin = datetime.strptime(demande.date_fin, "%Y-%m-%d")
    nombre_jours = (date_fin - date_debut).days + 1
    
    demande_obj = DemandeCongé(
        **demande.dict(),
        demandeur_id=current_user.id,
        nombre_jours=nombre_jours
    )
    await db.demandes_conge.insert_one(demande_obj.dict())
    
    # Créer notification pour approbation
    if current_user.role == "employe":
        # Notifier les superviseurs et admins
        superviseurs_admins = await db.users.find({"role": {"$in": ["superviseur", "admin"]}}).to_list(100)
        for superviseur in superviseurs_admins:
            await creer_notification(
                destinataire_id=superviseur["id"],
                type="conge_demande",
                titre="Nouvelle demande de congé",
                message=f"{current_user.prenom} {current_user.nom} demande un congé ({demande.type_conge}) du {demande.date_debut} au {demande.date_fin}",
                lien="/conges",
                data={"demande_id": demande_obj.id}
            )
    
    return demande_obj

@api_router.get("/demandes-conge", response_model=List[DemandeCongé])
async def get_demandes_conge(current_user: User = Depends(get_current_user)):
    if current_user.role == "employe":
        # Employés voient seulement leurs demandes
        demandes = await db.demandes_conge.find({"demandeur_id": current_user.id}).to_list(1000)
    else:
        # Superviseurs et admins voient toutes les demandes
        demandes = await db.demandes_conge.find().to_list(1000)
    
    cleaned_demandes = [clean_mongo_doc(demande) for demande in demandes]
    return [DemandeCongé(**demande) for demande in cleaned_demandes]

@api_router.put("/demandes-conge/{demande_id}/approuver")
async def approuver_demande_conge(demande_id: str, action: str, commentaire: str = "", current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    demande = await db.demandes_conge.find_one({"id": demande_id})
    if not demande:
        raise HTTPException(status_code=404, detail="Demande non trouvée")
    
    # Vérifier les permissions : superviseur peut approuver employés, admin peut tout approuver
    demandeur = await db.users.find_one({"id": demande["demandeur_id"]})
    if current_user.role == "superviseur" and demandeur["role"] != "employe":
        raise HTTPException(status_code=403, detail="Un superviseur ne peut approuver que les demandes d'employés")
    
    statut = "approuve" if action == "approuver" else "refuse"
    
    await db.demandes_conge.update_one(
        {"id": demande_id},
        {
            "$set": {
                "statut": statut,
                "approuve_par": current_user.id,
                "date_approbation": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
                "commentaire_approbation": commentaire
            }
        }
    )
    
    # Créer notification pour le demandeur
    demandeur = await db.users.find_one({"id": demande["demandeur_id"]})
    if demandeur:
        titre = f"Congé {statut}" if statut == "approuve" else "Congé refusé"
        message = f"Votre demande de congé du {demande['date_debut']} au {demande['date_fin']} a été {statut}e"
        if commentaire:
            message += f". Commentaire: {commentaire}"
        
        await creer_notification(
            destinataire_id=demande["demandeur_id"],
            type=f"conge_{statut}",
            titre=titre,
            message=message,
            lien="/conges",
            data={"demande_id": demande_id}
        )
    
    return {"message": f"Demande {statut}e avec succès"}

# Algorithme intelligent de recherche de remplaçants
@api_router.post("/remplacements/{demande_id}/recherche-automatique")
async def recherche_remplacants_automatique(demande_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Récupérer la demande de remplacement
        demande = await db.demandes_remplacement.find_one({"id": demande_id})
        if not demande:
            raise HTTPException(status_code=404, detail="Demande non trouvée")
        
        # Récupérer les paramètres de remplacement
        # (selon les règles définies dans Paramètres > Remplacements)
        
        # Trouver les remplaçants potentiels selon l'algorithme intelligent
        users = await db.users.find({"statut": "Actif"}).to_list(1000)
        type_garde = await db.types_garde.find_one({"id": demande["type_garde_id"]})
        
        remplacants_potentiels = []
        
        for user in users:
            if user["id"] == demande["demandeur_id"]:
                continue  # Skip demandeur
                
            # Étape 1: Vérifier disponibilités (si temps partiel)
            if user["type_emploi"] == "temps_partiel":
                # Get user disponibilités pour cette date exacte
                user_dispos = await db.disponibilites.find({
                    "user_id": user["id"],
                    "date": demande["date"],
                    "statut": "disponible"
                }).to_list(10)
                
                # Vérifier si disponible pour ce type de garde spécifiquement
                type_garde_compatible = any(
                    d.get("type_garde_id") == type_garde["id"] or d.get("type_garde_id") is None 
                    for d in user_dispos
                )
                
                if not type_garde_compatible:
                    continue  # Skip si pas disponible pour ce type de garde
            
            # Étape 2: Vérifier grade équivalent (si paramètre activé)
            # Étape 3: Vérifier compétences équivalentes (si paramètre activé)
            
            remplacants_potentiels.append({
                "user_id": user["id"],
                "nom": f"{user['prenom']} {user['nom']}",
                "grade": user["grade"],
                "score_compatibilite": 85  # Algorithme de scoring à développer
            })
        
        # Trier par score de compatibilité
        remplacants_potentiels.sort(key=lambda x: x["score_compatibilite"], reverse=True)
        
        # Limiter selon max_personnes_contact des paramètres
        max_contacts = 5  # À récupérer des paramètres
        remplacants_finaux = remplacants_potentiels[:max_contacts]
        
        # Créer les notifications pour les remplaçants potentiels
        for remplacant in remplacants_finaux:
            notification = NotificationRemplacement(
                demande_remplacement_id=demande_id,
                destinataire_id=remplacant["user_id"],
                message=f"Remplacement disponible le {demande['date']} - {type_garde['nom'] if type_garde else 'Garde'}",
                type_notification="remplacement_disponible"
            )
            await db.notifications.insert_one(notification.dict())
        
        return {
            "message": "Recherche automatique effectuée",
            "remplacants_contactes": len(remplacants_finaux),
            "algorithme": "Disponibilités → Grade → Compétences → Score compatibilité"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur recherche automatique: {str(e)}")

# Rapports et exports routes
@api_router.get("/rapports/export-pdf")
async def export_pdf_report(type_rapport: str = "general", user_id: str = None, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # En-tête du rapport
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#dc2626')
        )
        
        story.append(Paragraph("ProFireManager v2.0 - Rapport d'Activité", title_style))
        story.append(Spacer(1, 12))
        
        if type_rapport == "general":
            # Rapport général
            story.append(Paragraph("📊 Statistiques Générales", styles['Heading2']))
            
            # Récupérer les données
            users = await db.users.find({"statut": "Actif"}).to_list(1000)
            assignations = await db.assignations.find().to_list(1000)
            formations = await db.formations.find().to_list(1000)
            
            data = [
                ['Indicateur', 'Valeur'],
                ['Personnel actif', str(len(users))],
                ['Assignations totales', str(len(assignations))],
                ['Formations disponibles', str(len(formations))],
                ['Employés temps plein', str(len([u for u in users if u.get('type_emploi') == 'temps_plein']))],
                ['Employés temps partiel', str(len([u for u in users if u.get('type_emploi') == 'temps_partiel']))],
            ]
            
            table = Table(data)
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            
            story.append(table)
            
        elif type_rapport == "employe" and user_id:
            # Rapport par employé
            user_data = await db.users.find_one({"id": user_id})
            if user_data:
                story.append(Paragraph(f"👤 Rapport Personnel - {user_data['prenom']} {user_data['nom']}", styles['Heading2']))
                
                user_assignations = await db.assignations.find({"user_id": user_id}).to_list(1000)
                
                data = [
                    ['Information', 'Détail'],
                    ['Nom complet', f"{user_data['prenom']} {user_data['nom']}"],
                    ['Grade', user_data['grade']],
                    ['Type emploi', user_data['type_emploi']],
                    ['Gardes assignées', str(len(user_assignations))],
                    ['Statut', user_data['statut']]
                ]
                
                table = Table(data)
                table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#dc2626')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(table)
        
        doc.build(story)
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Retourner en base64 pour le frontend
        pdf_base64 = base64.b64encode(pdf_data).decode('utf-8')
        
        return {
            "message": "Rapport PDF généré avec succès",
            "filename": f"rapport_{type_rapport}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
            "data": pdf_base64
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur génération PDF: {str(e)}")

@api_router.get("/rapports/export-excel")
async def export_excel_report(type_rapport: str = "general", current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        wb = Workbook()
        ws = wb.active
        
        # Style de l'en-tête
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        
        if type_rapport == "general":
            ws.title = "Rapport Général"
            
            # En-tête
            headers = ["Indicateur", "Valeur", "Détails"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Données
            users = await db.users.find({"statut": "Actif"}).to_list(1000)
            assignations = await db.assignations.find().to_list(1000)
            
            data_rows = [
                ["Personnel Total", len(users), f"{len([u for u in users if u.get('type_emploi') == 'temps_plein'])} temps plein, {len([u for u in users if u.get('type_emploi') == 'temps_partiel'])} temps partiel"],
                ["Assignations", len(assignations), f"Période: {datetime.now().strftime('%B %Y')}"],
                ["Taux Activité", "85%", "Personnel actif vs total"],
            ]
            
            for row, (indicateur, valeur, details) in enumerate(data_rows, 2):
                ws.cell(row=row, column=1, value=indicateur)
                ws.cell(row=row, column=2, value=valeur)
                ws.cell(row=row, column=3, value=details)
        
        # Sauvegarder en mémoire
        buffer = BytesIO()
        wb.save(buffer)
        excel_data = buffer.getvalue()
        buffer.close()
        
        # Retourner en base64
        excel_base64 = base64.b64encode(excel_data).decode('utf-8')
        
        return {
            "message": "Rapport Excel généré avec succès",
            "filename": f"rapport_{type_rapport}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "data": excel_base64
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur génération Excel: {str(e)}")

@api_router.get("/rapports/statistiques-avancees")
async def get_statistiques_avancees(current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Récupérer toutes les données nécessaires
        users = await db.users.find().to_list(1000)
        assignations = await db.assignations.find().to_list(1000)
        types_garde = await db.types_garde.find().to_list(1000)
        formations = await db.formations.find().to_list(1000)
        demandes_remplacement = await db.demandes_remplacement.find().to_list(1000)
        
        # Statistiques générales
        stats_generales = {
            "personnel_total": len(users),
            "personnel_actif": len([u for u in users if u.get("statut") == "Actif"]),
            "assignations_mois": len(assignations),
            "taux_couverture": 94.5,  # Calcul à améliorer
            "formations_disponibles": len(formations),
            "remplacements_demandes": len(demandes_remplacement)
        }
        
        # Statistiques par rôle
        stats_par_role = {}
        for role in ["admin", "superviseur", "employe"]:
            users_role = [u for u in users if u.get("role") == role]
            assignations_role = [a for a in assignations if any(u["id"] == a["user_id"] and u.get("role") == role for u in users)]
            
            stats_par_role[role] = {
                "nombre_utilisateurs": len(users_role),
                "assignations_totales": len(assignations_role),
                "heures_moyennes": len(assignations_role) * 8,  # Estimation
                "formations_completees": sum(len(u.get("formations", [])) for u in users_role)
            }
        
        # Statistiques par employé (pour export individuel)
        stats_par_employe = []
        for user in users:
            user_assignations = [a for a in assignations if a["user_id"] == user["id"]]
            user_disponibilites = await db.disponibilites.find({"user_id": user["id"]}).to_list(100)
            
            stats_par_employe.append({
                "id": user["id"],
                "nom": f"{user['prenom']} {user['nom']}",
                "grade": user["grade"],
                "role": user["role"],
                "type_emploi": user["type_emploi"],
                "assignations_count": len(user_assignations),
                "disponibilites_count": len(user_disponibilites),
                "formations_count": len(user.get("formations", [])),
                "heures_estimees": len(user_assignations) * 8
            })
        
        return {
            "statistiques_generales": stats_generales,
            "statistiques_par_role": stats_par_role,
            "statistiques_par_employe": stats_par_employe,
            "periode": datetime.now().strftime("%B %Y"),
            "date_generation": datetime.now().isoformat()
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur calcul statistiques: {str(e)}")

# Sessions de formation routes
@api_router.post("/sessions-formation", response_model=SessionFormation)
async def create_session_formation(session: SessionFormationCreate, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    session_obj = SessionFormation(**session.dict())
    await db.sessions_formation.insert_one(session_obj.dict())
    return session_obj

@api_router.get("/sessions-formation", response_model=List[SessionFormation])
async def get_sessions_formation(current_user: User = Depends(get_current_user)):
    sessions = await db.sessions_formation.find().to_list(1000)
    cleaned_sessions = [clean_mongo_doc(session) for session in sessions]
    return [SessionFormation(**session) for session in cleaned_sessions]

@api_router.post("/sessions-formation/{session_id}/inscription")
async def inscrire_formation(session_id: str, current_user: User = Depends(get_current_user)):
    # Vérifier que la session existe
    session = await db.sessions_formation.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session de formation non trouvée")
    
    # Vérifier si déjà inscrit
    if current_user.id in session.get("participants", []):
        raise HTTPException(status_code=400, detail="Vous êtes déjà inscrit à cette formation")
    
    # Vérifier les places disponibles
    if len(session.get("participants", [])) >= session.get("places_max", 20):
        raise HTTPException(status_code=400, detail="Formation complète - Plus de places disponibles")
    
    # Ajouter l'utilisateur aux participants
    await db.sessions_formation.update_one(
        {"id": session_id},
        {"$push": {"participants": current_user.id}}
    )
    
    # Créer l'inscription
    inscription_obj = InscriptionFormation(
        session_id=session_id,
        user_id=current_user.id
    )
    await db.inscriptions_formation.insert_one(inscription_obj.dict())
    
    return {"message": "Inscription réussie", "session_id": session_id}

@api_router.delete("/sessions-formation/{session_id}/desinscription")
async def desinscrire_formation(session_id: str, current_user: User = Depends(get_current_user)):
    # Vérifier que la session existe
    session = await db.sessions_formation.find_one({"id": session_id})
    if not session:
        raise HTTPException(status_code=404, detail="Session de formation non trouvée")
    
    # Vérifier si inscrit
    if current_user.id not in session.get("participants", []):
        raise HTTPException(status_code=400, detail="Vous n'êtes pas inscrit à cette formation")
    
    # Retirer l'utilisateur des participants
    await db.sessions_formation.update_one(
        {"id": session_id},
        {"$pull": {"participants": current_user.id}}
    )
    
    # Supprimer l'inscription
    await db.inscriptions_formation.delete_one({
        "session_id": session_id,
        "user_id": current_user.id
    })
    
    return {"message": "Désinscription réussie", "session_id": session_id}

# Disponibilités routes
@api_router.post("/disponibilites", response_model=Disponibilite)
async def create_disponibilite(disponibilite: DisponibiliteCreate, current_user: User = Depends(get_current_user)):
    disponibilite_obj = Disponibilite(**disponibilite.dict())
    await db.disponibilites.insert_one(disponibilite_obj.dict())
    return disponibilite_obj

@api_router.get("/disponibilites/{user_id}", response_model=List[Disponibilite])
async def get_user_disponibilites(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    disponibilites = await db.disponibilites.find({"user_id": user_id}).to_list(1000)
    cleaned_disponibilites = [clean_mongo_doc(dispo) for dispo in disponibilites]
    return [Disponibilite(**dispo) for dispo in cleaned_disponibilites]

@api_router.put("/disponibilites/{user_id}")
async def update_user_disponibilites(user_id: str, disponibilites: List[DisponibiliteCreate], current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Delete existing disponibilités for this user
    await db.disponibilites.delete_many({"user_id": user_id})
    
    # Insert new disponibilités
    if disponibilites:
        dispo_docs = []
        for dispo in disponibilites:
            dispo_obj = Disponibilite(**dispo.dict())
            dispo_docs.append(dispo_obj.dict())
        
        await db.disponibilites.insert_many(dispo_docs)
    
    return {"message": f"Disponibilités mises à jour avec succès ({len(disponibilites)} entrées)"}

@api_router.delete("/disponibilites/{disponibilite_id}")
async def delete_disponibilite(disponibilite_id: str, current_user: User = Depends(get_current_user)):
    # Find the disponibilité to check ownership
    disponibilite = await db.disponibilites.find_one({"id": disponibilite_id})
    if not disponibilite:
        raise HTTPException(status_code=404, detail="Disponibilité non trouvée")
    
    if current_user.role not in ["admin", "superviseur"] and current_user.id != disponibilite["user_id"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    result = await db.disponibilites.delete_one({"id": disponibilite_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=400, detail="Impossible de supprimer la disponibilité")
    
    return {"message": "Disponibilité supprimée avec succès"}

# Assignation manuelle avancée avec récurrence
@api_router.post("/planning/assignation-avancee")
async def assignation_manuelle_avancee(
    assignation_data: dict,
    current_user: User = Depends(get_current_user)
):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        user_id = assignation_data.get("user_id")
        type_garde_id = assignation_data.get("type_garde_id")
        recurrence_type = assignation_data.get("recurrence_type", "unique")
        date_debut = datetime.strptime(assignation_data.get("date_debut"), "%Y-%m-%d").date()
        date_fin = datetime.strptime(assignation_data.get("date_fin", assignation_data.get("date_debut")), "%Y-%m-%d").date()
        jours_semaine = assignation_data.get("jours_semaine", [])
        
        assignations_creees = []
        
        if recurrence_type == "unique":
            # Assignation unique
            assignation_obj = Assignation(
                user_id=user_id,
                type_garde_id=type_garde_id,
                date=date_debut.strftime("%Y-%m-%d"),
                assignation_type="manuel_avance"
            )
            await db.assignations.insert_one(assignation_obj.dict())
            assignations_creees.append(assignation_obj.dict())
            
        elif recurrence_type == "hebdomadaire":
            # Récurrence hebdomadaire
            current_date = date_debut
            jours_semaine_index = {
                'monday': 0, 'tuesday': 1, 'wednesday': 2, 'thursday': 3,
                'friday': 4, 'saturday': 5, 'sunday': 6
            }
            
            while current_date <= date_fin:
                day_name = current_date.strftime("%A").lower()
                
                if day_name in jours_semaine:
                    # Vérifier qu'il n'y a pas déjà une assignation
                    existing = await db.assignations.find_one({
                        "user_id": user_id,
                        "type_garde_id": type_garde_id,
                        "date": current_date.strftime("%Y-%m-%d")
                    })
                    
                    if not existing:
                        assignation_obj = Assignation(
                            user_id=user_id,
                            type_garde_id=type_garde_id,
                            date=current_date.strftime("%Y-%m-%d"),
                            assignation_type="manuel_avance"
                        )
                        await db.assignations.insert_one(assignation_obj.dict())
                        assignations_creees.append(assignation_obj.dict())
                
                current_date += timedelta(days=1)
                
        elif recurrence_type == "mensuel":
            # Récurrence mensuelle (même jour du mois)
            jour_mois = date_debut.day
            current_month = date_debut.replace(day=1)
            
            while current_month <= date_fin:
                try:
                    # Essayer de créer la date pour ce mois
                    target_date = current_month.replace(day=jour_mois)
                    
                    if date_debut <= target_date <= date_fin:
                        existing = await db.assignations.find_one({
                            "user_id": user_id,
                            "type_garde_id": type_garde_id,
                            "date": target_date.strftime("%Y-%m-%d")
                        })
                        
                        if not existing:
                            assignation_obj = Assignation(
                                user_id=user_id,
                                type_garde_id=type_garde_id,
                                date=target_date.strftime("%Y-%m-%d"),
                                assignation_type="manuel_avance"
                            )
                            await db.assignations.insert_one(assignation_obj.dict())
                            assignations_creees.append(assignation_obj.dict())
                            
                except ValueError:
                    # Jour n'existe pas dans ce mois (ex: 31 février)
                    pass
                
                # Passer au mois suivant
                if current_month.month == 12:
                    current_month = current_month.replace(year=current_month.year + 1, month=1)
                else:
                    current_month = current_month.replace(month=current_month.month + 1)
        
        return {
            "message": "Assignation avancée créée avec succès",
            "assignations_creees": len(assignations_creees),
            "recurrence": recurrence_type,
            "periode": f"{date_debut.strftime('%Y-%m-%d')} à {date_fin.strftime('%Y-%m-%d')}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur assignation avancée: {str(e)}")

# Mode démo spécial - Attribution automatique agressive pour impression client
@api_router.post("/planning/attribution-auto-demo")
async def attribution_automatique_demo(semaine_debut: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Get all available users and types de garde
        users = await db.users.find({"statut": "Actif"}).to_list(1000)
        types_garde = await db.types_garde.find().to_list(1000)
        
        # Get existing assignations for the week
        semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        existing_assignations = await db.assignations.find({
            "date": {
                "$gte": semaine_debut,
                "$lte": semaine_fin
            }
        }).to_list(1000)
        
        nouvelles_assignations = []
        
        # MODE DÉMO AGRESSIF - REMPLIR AU MAXIMUM
        for type_garde in types_garde:
            for day_offset in range(7):
                current_date = datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=day_offset)
                date_str = current_date.strftime("%Y-%m-%d")
                day_name = current_date.strftime("%A").lower()
                
                # Skip if type garde doesn't apply to this day
                if type_garde.get("jours_application") and day_name not in type_garde["jours_application"]:
                    continue
                
                # Compter combien de personnel déjà assigné pour cette garde
                existing_for_garde = [a for a in existing_assignations 
                                    if a["date"] == date_str and a["type_garde_id"] == type_garde["id"]]
                
                personnel_deja_assigne = len(existing_for_garde)
                personnel_requis = type_garde.get("personnel_requis", 1)
                
                # Assigner jusqu'au maximum requis
                for i in range(personnel_requis - personnel_deja_assigne):
                    # Trouver utilisateurs disponibles
                    available_users = []
                    
                    for user in users:
                        # Skip si déjà assigné cette garde ce jour
                        if any(a["user_id"] == user["id"] and a["date"] == date_str and a["type_garde_id"] == type_garde["id"] 
                               for a in existing_assignations):
                            continue
                        
                        # Skip si déjà assigné autre garde ce jour (éviter conflits)
                        if any(a["user_id"] == user["id"] and a["date"] == date_str 
                               for a in existing_assignations):
                            continue
                        
                        # Vérifier disponibilités
                        user_dispos = await db.disponibilites.find({
                            "user_id": user["id"],
                            "date": date_str,
                            "type_garde_id": type_garde["id"],
                            "statut": "disponible"
                        }).to_list(10)
                        
                        if user_dispos:
                            available_users.append(user)
                    
                    if not available_users:
                        break  # Pas d'utilisateurs disponibles pour ce poste
                    
                    # MODE DÉMO : ASSOUPLIR CONTRAINTE OFFICIER
                    if type_garde.get("officier_obligatoire", False):
                        # Chercher officiers d'abord
                        officers = [u for u in available_users if u["grade"] in ["Capitaine", "Lieutenant", "Directeur"]]
                        # Sinon pompiers avec fonction supérieur
                        if not officers:
                            officers = [u for u in available_users if u.get("fonction_superieur", False)]
                        # En dernier recours : tous pompiers (MODE DÉMO)
                        if not officers:
                            officers = available_users
                        
                        if officers:
                            selected_user = officers[0]
                        else:
                            continue
                    else:
                        selected_user = available_users[0]
                    
                    # Créer assignation
                    assignation_obj = Assignation(
                        user_id=selected_user["id"],
                        type_garde_id=type_garde["id"],
                        date=date_str,
                        assignation_type="auto_demo"
                    )
                    
                    await db.assignations.insert_one(assignation_obj.dict())
                    nouvelles_assignations.append(assignation_obj.dict())
                    existing_assignations.append(assignation_obj.dict())
        
        return {
            "message": "Attribution DÉMO agressive effectuée avec succès",
            "assignations_creees": len(nouvelles_assignations),
            "algorithme": "Mode démo : Contraintes assouplies pour impression maximum",
            "semaine": f"{semaine_debut} - {semaine_fin}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur attribution démo: {str(e)}")

# Attribution automatique intelligente avec rotation équitable et ancienneté
@api_router.post("/planning/attribution-auto")
async def attribution_automatique(semaine_debut: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Get all available users and types de garde
        users = await db.users.find({"statut": "Actif"}).to_list(1000)
        types_garde = await db.types_garde.find().to_list(1000)
        
        # Get existing assignations for the week
        semaine_fin = (datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=6)).strftime("%Y-%m-%d")
        existing_assignations = await db.assignations.find({
            "date": {
                "$gte": semaine_debut,
                "$lte": semaine_fin
            }
        }).to_list(1000)
        
        # Get monthly statistics for rotation équitable (current month)
        current_month_start = datetime.strptime(semaine_debut, "%Y-%m-%d").replace(day=1).strftime("%Y-%m-%d")
        current_month_end = (datetime.strptime(current_month_start, "%Y-%m-%d") + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        current_month_end = current_month_end.strftime("%Y-%m-%d")
        
        monthly_assignations = await db.assignations.find({
            "date": {
                "$gte": current_month_start,
                "$lte": current_month_end
            }
        }).to_list(1000)
        
        # Calculate monthly hours for each user
        user_monthly_hours = {}
        for user in users:
            user_hours = 0
            for assignation in monthly_assignations:
                if assignation["user_id"] == user["id"]:
                    # Find type garde to get duration
                    type_garde = next((t for t in types_garde if t["id"] == assignation["type_garde_id"]), None)
                    if type_garde:
                        user_hours += type_garde.get("duree_heures", 8)
            user_monthly_hours[user["id"]] = user_hours
        
        # Attribution automatique logic (5 niveaux de priorité)
        nouvelles_assignations = []
        
        for type_garde in types_garde:
            # Check each day for this type de garde
            for day_offset in range(7):
                current_date = datetime.strptime(semaine_debut, "%Y-%m-%d") + timedelta(days=day_offset)
                date_str = current_date.strftime("%Y-%m-%d")
                day_name = current_date.strftime("%A").lower()
                
                # Skip if type garde doesn't apply to this day
                if type_garde.get("jours_application") and day_name not in type_garde["jours_application"]:
                    continue
                
                # ÉTAPE 1: Check if already assigned manually
                existing = next((a for a in existing_assignations 
                               if a["date"] == date_str and a["type_garde_id"] == type_garde["id"]), None)
                if existing and existing.get("assignation_type") == "manuel":
                    continue  # Respecter les assignations manuelles
                
                # Find available users for this slot
                available_users = []
                for user in users:
                    # ÉTAPE 2: Check if user has availability (for part-time employees)
                    if user["type_emploi"] == "temps_partiel":
                        # Get user disponibilités
                        user_dispos = await db.disponibilites.find({
                            "user_id": user["id"],
                            "date": date_str,
                            "statut": "disponible"
                        }).to_list(10)
                        
                        if not user_dispos:
                            continue  # Skip if not available
                    else:
                        # Skip temps plein (planning fixe manuel)
                        continue
                    
                    # Check if user already assigned on this date
                    already_assigned = next((a for a in existing_assignations 
                                           if a["date"] == date_str and a["user_id"] == user["id"]), None)
                    if already_assigned:
                        continue
                    
                    available_users.append(user)
                
                if not available_users:
                    continue
                
                # ÉTAPE 3: Apply grade requirements (1 officier obligatoire si configuré)
                if type_garde.get("officier_obligatoire", False):
                    # Filter officers (Capitaine, Lieutenant, Directeur)
                    officers = [u for u in available_users if u["grade"] in ["Capitaine", "Lieutenant", "Directeur"]]
                    if officers:
                        available_users = officers
                
                # ÉTAPE 4: Rotation équitable - sort by monthly hours (ascending)
                available_users.sort(key=lambda u: user_monthly_hours.get(u["id"], 0))
                
                # ÉTAPE 5: Ancienneté - among users with same hours, prioritize by ancienneté
                min_hours = user_monthly_hours.get(available_users[0]["id"], 0)
                users_with_min_hours = [u for u in available_users if user_monthly_hours.get(u["id"], 0) == min_hours]
                
                if len(users_with_min_hours) > 1:
                    # Sort by ancienneté (date_embauche) - oldest first
                    users_with_min_hours.sort(key=lambda u: datetime.strptime(u["date_embauche"], "%d/%m/%Y"))
                
                # Select the best candidate
                selected_user = users_with_min_hours[0]
                
                assignation_obj = Assignation(
                    user_id=selected_user["id"],
                    type_garde_id=type_garde["id"],
                    date=date_str,
                    assignation_type="auto"
                )
                
                await db.assignations.insert_one(assignation_obj.dict())
                nouvelles_assignations.append(assignation_obj.dict())
                existing_assignations.append(assignation_obj.dict())
                
                # Update monthly hours for next iteration
                user_monthly_hours[selected_user["id"]] += type_garde.get("duree_heures", 8)
        
        return {
            "message": "Attribution automatique intelligente effectuée avec succès",
            "assignations_creees": len(nouvelles_assignations),
            "algorithme": "5 niveaux: Manuel → Disponibilités → Grades → Rotation équitable → Ancienneté",
            "semaine": f"{semaine_debut} - {semaine_fin}"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors de l'attribution automatique: {str(e)}")

# Endpoint pour obtenir les statistiques personnelles mensuelles
@api_router.get("/users/{user_id}/stats-mensuelles")
async def get_user_monthly_stats(user_id: str, current_user: User = Depends(get_current_user)):
    if current_user.role not in ["admin", "superviseur"] and current_user.id != user_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Get current month assignations for this user
        today = datetime.now(timezone.utc)
        month_start = today.replace(day=1).strftime("%Y-%m-%d")
        month_end = (today.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        month_end = month_end.strftime("%Y-%m-%d")
        
        user_assignations = await db.assignations.find({
            "user_id": user_id,
            "date": {
                "$gte": month_start,
                "$lte": month_end
            }
        }).to_list(1000)
        
        # Get types garde for calculating hours
        types_garde = await db.types_garde.find().to_list(1000)
        types_dict = {t["id"]: t for t in types_garde}
        
        # Calculate stats
        gardes_ce_mois = len(user_assignations)
        heures_travaillees = 0
        
        for assignation in user_assignations:
            type_garde = types_dict.get(assignation["type_garde_id"])
            if type_garde:
                heures_travaillees += type_garde.get("duree_heures", 8)
        
        # Get user formations count
        user_data = await db.users.find_one({"id": user_id})
        certifications = len(user_data.get("formations", [])) if user_data else 0
        
        return {
            "gardes_ce_mois": gardes_ce_mois,
            "heures_travaillees": heures_travaillees,
            "certifications": certifications,
            "mois": today.strftime("%B %Y")
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors du calcul des statistiques: {str(e)}")

# Statistics routes
@api_router.get("/statistiques", response_model=Statistiques)
async def get_statistiques(current_user: User = Depends(get_current_user)):
    try:
        # 1. Personnel actif (100% dynamique)
        personnel_count = await db.users.count_documents({"statut": "Actif"})
        
        # 2. Gardes cette semaine (100% dynamique)
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        gardes_count = await db.assignations.count_documents({
            "date": {
                "$gte": start_week.strftime("%Y-%m-%d"),
                "$lte": end_week.strftime("%Y-%m-%d")
            }
        })
        
        # 3. Formations planifiées (100% dynamique)
        formations_count = await db.sessions_formation.count_documents({"statut": "planifie"})
        
        # 4. Taux de couverture dynamique - CALCUL CORRECT
        # Calculer le total de personnel requis pour la semaine
        total_assignations_required = await db.types_garde.find().to_list(1000)
        total_personnel_requis = 0
        total_personnel_assigne = 0
        
        # Pour chaque jour de la semaine
        for day_offset in range(7):
            current_day = start_week + timedelta(days=day_offset)
            day_name = current_day.strftime("%A").lower()
            
            # Pour chaque type de garde
            for type_garde in total_assignations_required:
                jours_app = type_garde.get("jours_application", [])
                
                # Si ce type de garde s'applique à ce jour
                if not jours_app or day_name in jours_app:
                    personnel_requis = type_garde.get("personnel_requis", 1)
                    total_personnel_requis += personnel_requis
                    
                    # Compter combien de personnes sont assignées pour cette garde ce jour
                    assignations_jour = await db.assignations.count_documents({
                        "date": current_day.strftime("%Y-%m-%d"),
                        "type_garde_id": type_garde["id"]
                    })
                    
                    total_personnel_assigne += min(assignations_jour, personnel_requis)
        
        # Calcul correct : (personnel assigné / personnel requis) × 100
        taux_couverture = (total_personnel_assigne / total_personnel_requis * 100) if total_personnel_requis > 0 else 0
        
        # Cap à 100% maximum
        taux_couverture = min(taux_couverture, 100.0)
        
        # 5. Heures travaillées ce mois (100% dynamique)
        start_month = today.replace(day=1)
        end_month = (start_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        assignations_mois = await db.assignations.find({
            "date": {
                "$gte": start_month.strftime("%Y-%m-%d"),
                "$lte": end_month.strftime("%Y-%m-%d")
            }
        }).to_list(1000)
        
        # Calculer les heures basées sur les types de garde
        heures_totales = 0
        types_garde_dict = {tg["id"]: tg for tg in total_assignations_required}
        
        for assignation in assignations_mois:
            type_garde = types_garde_dict.get(assignation["type_garde_id"])
            if type_garde:
                heures_totales += type_garde.get("duree_heures", 8)
        
        # 6. Remplacements effectués (100% dynamique)
        remplacements_count = await db.demandes_remplacement.count_documents({"statut": "approuve"})
        
        return Statistiques(
            personnel_actif=personnel_count,
            gardes_cette_semaine=gardes_count,
            formations_planifiees=formations_count,
            taux_couverture=round(taux_couverture, 1),
            heures_travaillees=heures_totales,
            remplacements_effectues=remplacements_count
        )
        
    except Exception as e:
        # Fallback en cas d'erreur
        print(f"Erreur calcul statistiques: {str(e)}")
        return Statistiques(
            personnel_actif=0,
            gardes_cette_semaine=0,
            formations_planifiees=0,
            taux_couverture=0.0,
            heures_travaillees=0,
            remplacements_effectues=0
        )

# Réinitialiser tout le planning (vider toutes assignations)
@api_router.post("/planning/reinitialiser")
async def reinitialiser_planning(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Supprimer toutes les assignations
        result = await db.assignations.delete_many({})
        
        return {
            "message": "Planning réinitialisé avec succès",
            "assignations_supprimees": result.deleted_count
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur réinitialisation: {str(e)}")

# Réparer tous les mots de passe démo
@api_router.post("/repair-demo-passwords")
async def repair_demo_passwords():
    try:
        password_fixes = [
            ("admin@firemanager.ca", "admin123"),
            ("superviseur@firemanager.ca", "superviseur123"),
            ("employe@firemanager.ca", "employe123"),
            ("partiel@firemanager.ca", "partiel123")
        ]
        
        fixed_count = 0
        for email, password in password_fixes:
            user = await db.users.find_one({"email": email})
            if user:
                new_hash = get_password_hash(password)
                await db.users.update_one(
                    {"email": email},
                    {"$set": {"mot_de_passe_hash": new_hash}}
                )
                fixed_count += 1
                print(f"Fixed password for {email}")
        
        return {"message": f"{fixed_count} mots de passe démo réparés"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Fix all demo passwords endpoint
@api_router.post("/fix-all-passwords")
async def fix_all_passwords():
    try:
        # Fix all demo account passwords
        password_fixes = [
            ("admin@firemanager.ca", "admin123"),
            ("superviseur@firemanager.ca", "superviseur123"),
            ("employe@firemanager.ca", "employe123"),
            ("partiel@firemanager.ca", "partiel123")
        ]
        
        fixed_count = 0
        for email, password in password_fixes:
            user = await db.users.find_one({"email": email})
            if user:
                new_hash = get_password_hash(password)
                await db.users.update_one(
                    {"email": email},
                    {"$set": {"mot_de_passe_hash": new_hash}}
                )
                fixed_count += 1
                print(f"Fixed password for {email}")
        
        return {"message": f"{fixed_count} mots de passe réparés"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Fix admin password endpoint
@api_router.post("/fix-admin-password")
async def fix_admin_password():
    try:
        # Find admin user
        admin_user = await db.users.find_one({"email": "admin@firemanager.ca"})
        if admin_user:
            # Update password hash
            new_password_hash = get_password_hash("admin123")
            await db.users.update_one(
                {"email": "admin@firemanager.ca"},
                {"$set": {"mot_de_passe_hash": new_password_hash}}
            )
            return {"message": "Mot de passe admin réparé"}
        else:
            return {"message": "Compte admin non trouvé"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Clean up endpoint
@api_router.post("/cleanup-duplicates")
async def cleanup_duplicates(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Clean formations duplicates - keep only unique ones by name
        formations = await db.formations.find().to_list(1000)
        unique_formations = {}
        
        for formation in formations:
            name = formation['nom']
            if name not in unique_formations:
                unique_formations[name] = formation
        
        # Delete all formations and re-insert unique ones
        await db.formations.delete_many({})
        
        if unique_formations:
            formations_to_insert = []
            for formation in unique_formations.values():
                formation.pop('_id', None)  # Remove MongoDB _id
                formations_to_insert.append(formation)
            
            await db.formations.insert_many(formations_to_insert)
        
        # Clean types garde duplicates
        types_garde = await db.types_garde.find().to_list(1000)
        unique_types = {}
        
        for type_garde in types_garde:
            key = f"{type_garde['nom']}_{type_garde['heure_debut']}_{type_garde['heure_fin']}"
            if key not in unique_types:
                unique_types[key] = type_garde
        
        # Delete all types garde and re-insert unique ones
        await db.types_garde.delete_many({})
        
        if unique_types:
            types_to_insert = []
            for type_garde in unique_types.values():
                type_garde.pop('_id', None)  # Remove MongoDB _id
                types_to_insert.append(type_garde)
            
            await db.types_garde.insert_many(types_to_insert)
        
        formations_count = len(unique_formations)
        types_count = len(unique_types)
        
        return {
            "message": f"Nettoyage terminé: {formations_count} formations uniques, {types_count} types de garde uniques"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur lors du nettoyage: {str(e)}")

# Créer données de démonstration réalistes avec historique
@api_router.post("/init-demo-data-realiste")
async def init_demo_data_realiste():
    try:
        # Clear existing data
        await db.users.delete_many({})
        await db.types_garde.delete_many({})
        await db.assignations.delete_many({})
        await db.planning.delete_many({})
        await db.demandes_remplacement.delete_many({})
        await db.formations.delete_many({})
        await db.sessions_formation.delete_many({})
        await db.disponibilites.delete_many({})
        
        # Créer plus d'utilisateurs réalistes
        demo_users = [
            {
                "nom": "Dupont", "prenom": "Jean", "email": "admin@firemanager.ca",
                "telephone": "514-111-2233", "contact_urgence": "514-999-1111",
                "grade": "Directeur", "fonction_superieur": False, "type_emploi": "temps_plein",
                "heures_max_semaine": 40, "role": "admin", "numero_employe": "ADM001",
                "date_embauche": "14/01/2020", "formations": [], "mot_de_passe": "admin123"
            },
            {
                "nom": "Dubois", "prenom": "Sophie", "email": "superviseur@firemanager.ca",
                "telephone": "514-444-5566", "contact_urgence": "514-888-2222",
                "grade": "Directeur", "fonction_superieur": False, "type_emploi": "temps_plein",
                "heures_max_semaine": 40, "role": "superviseur", "numero_employe": "POM001",
                "date_embauche": "07/01/2022", "formations": [], "mot_de_passe": "superviseur123"
            },
            {
                "nom": "Bernard", "prenom": "Pierre", "email": "employe@firemanager.ca",
                "telephone": "418-555-9999", "contact_urgence": "418-777-3333",
                "grade": "Capitaine", "fonction_superieur": False, "type_emploi": "temps_plein",
                "heures_max_semaine": 40, "role": "employe", "numero_employe": "POM002",
                "date_embauche": "21/09/2019", "formations": [], "mot_de_passe": "employe123"
            },
            {
                "nom": "Garcia", "prenom": "Claire", "email": "partiel@firemanager.ca",
                "telephone": "514-888-9900", "contact_urgence": "514-666-4444",
                "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel",
                "heures_max_semaine": 25, "role": "employe", "numero_employe": "POM005",
                "date_embauche": "02/11/2020", "formations": [], "mot_de_passe": "partiel123"
            },
            # Nouveaux utilisateurs pour démo réaliste
            {
                "nom": "Tremblay", "prenom": "Marc", "email": "marc.tremblay@firemanager.ca",
                "telephone": "418-222-3333", "contact_urgence": "418-999-4444",
                "grade": "Lieutenant", "fonction_superieur": False, "type_emploi": "temps_plein",
                "heures_max_semaine": 40, "role": "employe", "numero_employe": "POM003",
                "date_embauche": "15/03/2021", "formations": [], "mot_de_passe": "TempPass123!"
            },
            {
                "nom": "Martin", "prenom": "Sarah", "email": "sarah.martin@firemanager.ca",
                "telephone": "514-333-4444", "contact_urgence": "514-777-8888",
                "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_partiel",
                "heures_max_semaine": 20, "role": "employe", "numero_employe": "POM006",
                "date_embauche": "10/08/2023", "formations": [], "mot_de_passe": "TempPass123!"
            }
        ]
        
        # Créer formations avec plus de détails
        demo_formations = [
            {"nom": "Classe 4A", "description": "Formation de conduite véhicules lourds", "duree_heures": 40, "validite_mois": 60, "obligatoire": False},
            {"nom": "Désincarcération", "description": "Techniques de désincarcération", "duree_heures": 24, "validite_mois": 36, "obligatoire": True},
            {"nom": "Pompier 1", "description": "Formation de base pompier niveau 1", "duree_heures": 200, "validite_mois": 24, "obligatoire": True},
            {"nom": "Officier 2", "description": "Formation officier niveau 2", "duree_heures": 120, "validite_mois": 36, "obligatoire": False},
            {"nom": "Premiers Répondants", "description": "Formation premiers secours", "duree_heures": 16, "validite_mois": 12, "obligatoire": True},
            {"nom": "Sauvetage Aquatique", "description": "Techniques de sauvetage en milieu aquatique", "duree_heures": 32, "validite_mois": 24, "obligatoire": False}
        ]
        
        formation_ids = {}
        for formation_data in demo_formations:
            formation_obj = Formation(**formation_data)
            await db.formations.insert_one(formation_obj.dict())
            formation_ids[formation_data["nom"]] = formation_obj.id
        
        # Assigner formations aux utilisateurs
        demo_users[0]["formations"] = [formation_ids["Officier 2"], formation_ids["Pompier 1"]]
        demo_users[1]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"]]
        demo_users[2]["formations"] = [formation_ids["Classe 4A"], formation_ids["Désincarcération"], formation_ids["Premiers Répondants"]]
        demo_users[3]["formations"] = [formation_ids["Pompier 1"]]
        demo_users[4]["formations"] = [formation_ids["Désincarcération"], formation_ids["Premiers Répondants"], formation_ids["Sauvetage Aquatique"]]
        demo_users[5]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"]]
        
        # Créer utilisateurs
        user_ids = {}
        for user_data in demo_users:
            user_dict = user_data.copy()
            user_dict["mot_de_passe_hash"] = get_password_hash(user_dict.pop("mot_de_passe"))
            user_dict["statut"] = "Actif"
            user_obj = User(**user_dict)
            await db.users.insert_one(user_obj.dict())
            user_ids[user_data["email"]] = user_obj.id
        
        # Créer assignations historiques (3 mois)
        assignations_created = 0
        for week_offset in range(-12, 1):  # 12 semaines passées + semaine courante
            week_start = datetime.now(timezone.utc).date() + timedelta(weeks=week_offset)
            week_start = week_start - timedelta(days=week_start.weekday())
            
            for day_offset in range(7):
                date_assignation = week_start + timedelta(days=day_offset)
                date_str = date_assignation.strftime("%Y-%m-%d")
                
                # Assigner quelques gardes aléatoirement
                if assignations_created % 3 == 0:  # Environ 1/3 des jours
                    # Garde Interne AM
                    assignation_obj = Assignation(
                        user_id=user_ids["employe@firemanager.ca"],
                        type_garde_id="garde-interne-am",  # Sera créé après
                        date=date_str,
                        assignation_type="auto"
                    )
                    await db.assignations.insert_one(assignation_obj.dict())
                    assignations_created += 1
        
        return {"message": f"Données de démonstration réalistes créées : {len(demo_users)} utilisateurs, {len(demo_formations)} formations, {assignations_created} assignations historiques"}
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Affecter disponibilités à TOUS les employés temps partiel existants (auto-détection)
@api_router.post("/auto-affecter-disponibilites-temps-partiel")
async def auto_affecter_disponibilites_temps_partiel(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # DÉTECTION AUTOMATIQUE de tous les employés temps partiel
        tous_temps_partiel = await db.users.find({
            "type_emploi": "temps_partiel",
            "statut": "Actif"
        }).to_list(1000)
        
        print(f"Trouvé {len(tous_temps_partiel)} employés temps partiel")
        
        # Supprimer les anciennes disponibilités de la semaine courante
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        await db.disponibilites.delete_many({
            "date": {
                "$gte": start_week.strftime("%Y-%m-%d"),
                "$lte": end_week.strftime("%Y-%m-%d")
            }
        })
        
        # Récupérer types de garde
        types_garde = await db.types_garde.find().to_list(100)
        
        disponibilites_created = 0
        
        # AFFECTER DISPONIBILITÉS À TOUS LES TEMPS PARTIEL DÉTECTÉS
        for index, user in enumerate(tous_temps_partiel):
            print(f"Affectation pour {user['prenom']} {user['nom']} ({user['grade']})")
            
            # Pattern de disponibilité selon l'index pour variété
            if index % 4 == 0:  # Pattern 1: Lun-Mer-Ven
                jours_disponibles = [0, 2, 4]
            elif index % 4 == 1:  # Pattern 2: Mar-Jeu-Sam  
                jours_disponibles = [1, 3, 5]
            elif index % 4 == 2:  # Pattern 3: Mer-Ven-Dim
                jours_disponibles = [2, 4, 6]
            else:  # Pattern 4: Lun-Jeu-Dim
                jours_disponibles = [0, 3, 6]
            
            for day_offset in jours_disponibles:
                date_dispo = start_week + timedelta(days=day_offset)
                date_str = date_dispo.strftime("%Y-%m-%d")
                day_name = date_dispo.strftime("%A").lower()
                
                # Créer disponibilités pour TOUS les types de garde applicables
                for type_garde in types_garde:
                    jours_app = type_garde.get("jours_application", [])
                    if jours_app and day_name not in jours_app:
                        continue
                    
                    # Créer disponibilité (stratégie intensive pour démo)
                    dispo_obj = Disponibilite(
                        user_id=user["id"],
                        date=date_str,
                        type_garde_id=type_garde["id"],
                        heure_debut=type_garde["heure_debut"],
                        heure_fin=type_garde["heure_fin"],
                        statut="disponible"
                    )
                    await db.disponibilites.insert_one(dispo_obj.dict())
                    disponibilites_created += 1
        
        return {
            "message": "Disponibilités affectées automatiquement",
            "employes_temps_partiel_detectes": len(tous_temps_partiel),
            "disponibilites_creees": disponibilites_created,
            "semaine": f"{start_week.strftime('%Y-%m-%d')} - {end_week.strftime('%Y-%m-%d')}",
            "patterns": "4 patterns différents pour variété démo",
            "employés_détectés": [f"{u['prenom']} {u['nom']} ({u['grade']})" for u in tous_temps_partiel]
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur auto-affectation: {str(e)}")

# Créer disponibilités MAXIMALES pour démo parfaite
@api_router.post("/init-disponibilites-demo-complete")
async def init_disponibilites_demo_complete(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Supprimer toutes les disponibilités existantes
        await db.disponibilites.delete_many({})
        
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        # Récupérer TOUS les utilisateurs (temps plein ET temps partiel pour démo)
        all_users = await db.users.find({"statut": "Actif"}).to_list(100)
        types_garde = await db.types_garde.find().to_list(100)
        
        disponibilites_created = 0
        
        # STRATÉGIE DÉMO : TOUS LES EMPLOYÉS DISPONIBLES POUR TOUS LES TYPES
        for user in all_users:
            for day_offset in range(7):  # Chaque jour
                date_dispo = start_week + timedelta(days=day_offset)
                date_str = date_dispo.strftime("%Y-%m-%d")
                day_name = date_dispo.strftime("%A").lower()
                
                for type_garde in types_garde:
                    # Vérifier jours d'application
                    jours_app = type_garde.get("jours_application", [])
                    if jours_app and day_name not in jours_app:
                        continue
                    
                    # CRÉER DISPONIBILITÉ POUR TOUS (temps plein et temps partiel)
                    # Exception : respecter les heures max pour temps partiel
                    if user["type_emploi"] == "temps_partiel":
                        # Temps partiel : disponible seulement 3 jours par semaine
                        user_number = int(user["numero_employe"][-1]) if user["numero_employe"][-1].isdigit() else 0
                        
                        # Pattern par employé pour éviter épuisement
                        if user_number % 3 == 0 and day_offset in [0, 2, 4]:  # Lun-Mer-Ven
                            pass
                        elif user_number % 3 == 1 and day_offset in [1, 3, 5]:  # Mar-Jeu-Sam
                            pass  
                        elif user_number % 3 == 2 and day_offset in [2, 4, 6]:  # Mer-Ven-Dim
                            pass
                        else:
                            continue  # Skip ce jour pour cet employé
                    
                    # CRÉER DISPONIBILITÉ
                    dispo_obj = Disponibilite(
                        user_id=user["id"],
                        date=date_str,
                        type_garde_id=type_garde["id"],
                        heure_debut=type_garde["heure_debut"],
                        heure_fin=type_garde["heure_fin"],
                        statut="disponible"
                    )
                    await db.disponibilites.insert_one(dispo_obj.dict())
                    disponibilites_created += 1
        
        return {
            "message": "Disponibilités DÉMO COMPLÈTES créées",
            "semaine": f"{start_week.strftime('%Y-%m-%d')} - {end_week.strftime('%Y-%m-%d')}",
            "disponibilites_creees": disponibilites_created,
            "all_users_included": len(all_users),
            "strategy": "TOUS employés (TP+TPa) avec patterns optimisés"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Créer disponibilités pour semaine courante (démo assignation auto)
@api_router.post("/init-disponibilites-semaine-courante")
async def init_disponibilites_semaine_courante(current_user: User = Depends(get_current_user)):
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    try:
        # Supprimer les disponibilités existantes pour la semaine courante
        today = datetime.now(timezone.utc).date()
        start_week = today - timedelta(days=today.weekday())
        end_week = start_week + timedelta(days=6)
        
        await db.disponibilites.delete_many({
            "date": {
                "$gte": start_week.strftime("%Y-%m-%d"),
                "$lte": end_week.strftime("%Y-%m-%d")
            }
        })
        
        # Récupérer tous les types de garde
        types_garde = await db.types_garde.find().to_list(100)
        # DÉTECTION AUTOMATIQUE de TOUS les employés temps partiel (peu importe le nombre)
        tous_temps_partiel = await db.users.find({
            "type_emploi": "temps_partiel",
            "statut": "Actif"
        }).to_list(1000)
        
        print(f"AUTO-DÉTECTION: {len(tous_temps_partiel)} employés temps partiel trouvés")
        
        disponibilites_created = 0
        
        # ALGORITHME OPTIMISÉ POUR TOUS VOS EMPLOYÉS TEMPS PARTIEL
        for user_index, user in enumerate(tous_temps_partiel):
            for day_offset in range(7):  # Chaque jour de la semaine courante
                date_dispo = start_week + timedelta(days=day_offset)
                date_str = date_dispo.strftime("%Y-%m-%d")
                day_name = date_dispo.strftime("%A").lower()
                
                # Pattern de disponibilité varié selon l'employé
                if user_index % 3 == 0:  # 1/3 des employés : Lun-Mer-Ven
                    jours_pattern = ['monday', 'wednesday', 'friday']
                elif user_index % 3 == 1:  # 1/3 des employés : Mar-Jeu-Sam
                    jours_pattern = ['tuesday', 'thursday', 'saturday']
                else:  # 1/3 des employés : Mer-Ven-Dim
                    jours_pattern = ['wednesday', 'friday', 'sunday']
                
                if day_name in jours_pattern:
                    # Créer disponibilités pour TOUS les types de garde applicables
                    for type_garde in types_garde:
                        jours_app = type_garde.get("jours_application", [])
                        if jours_app and day_name not in jours_app:
                            continue
                        
                        # CRÉER DISPONIBILITÉ pour vos employés (pompiers ET lieutenants)
                        dispo_obj = Disponibilite(
                            user_id=user["id"],
                            date=date_str,
                            type_garde_id=type_garde["id"],
                            heure_debut=type_garde["heure_debut"],
                            heure_fin=type_garde["heure_fin"],
                            statut="disponible"
                        )
                        await db.disponibilites.insert_one(dispo_obj.dict())
                        disponibilites_created += 1
        
        return {
            "message": "Disponibilités créées pour TOUS vos employés temps partiel",
            "employes_temps_partiel": len(tous_temps_partiel),
            "disponibilites_creees": disponibilites_created,
            "all_users_included": len(tous_temps_partiel),
            "strategy": f"AUTO-DÉTECTION: {len(tous_temps_partiel)} employés temps partiel avec patterns optimisés"
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Créer données de démonstration OPTIMALES pour démo client
@api_router.post("/init-demo-client-data")
async def init_demo_client_data():
    try:
        # Clear existing data
        await db.users.delete_many({})
        await db.types_garde.delete_many({})
        await db.assignations.delete_many({})
        await db.formations.delete_many({})
        await db.sessions_formation.delete_many({})
        await db.disponibilites.delete_many({})
        await db.demandes_remplacement.delete_many({})
        
        # 1. ÉQUIPE RÉALISTE CASERNE (15 pompiers)
        demo_users = [
            # DIRECTION ET ADMINISTRATION
            {"nom": "Dupont", "prenom": "Jean", "email": "admin@firemanager.ca", "telephone": "514-111-2233", "contact_urgence": "514-999-1111", "grade": "Directeur", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "admin", "numero_employe": "DIR001", "date_embauche": "14/01/2015", "formations": [], "mot_de_passe": "admin123"},
            {"nom": "Tremblay", "prenom": "Marie", "email": "directrice@firemanager.ca", "telephone": "514-222-3344", "contact_urgence": "514-888-1111", "grade": "Directeur", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "superviseur", "numero_employe": "DIR002", "date_embauche": "03/06/2018", "formations": [], "mot_de_passe": "superviseur123"},
            
            # SUPERVISEURS / CAPITAINES
            {"nom": "Dubois", "prenom": "Sophie", "email": "superviseur@firemanager.ca", "telephone": "514-444-5566", "contact_urgence": "514-888-2222", "grade": "Capitaine", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "superviseur", "numero_employe": "CAP001", "date_embauche": "07/01/2019", "formations": [], "mot_de_passe": "superviseur123"},
            {"nom": "Leblanc", "prenom": "Michel", "email": "michel.leblanc@firemanager.ca", "telephone": "418-333-4455", "contact_urgence": "418-777-5555", "grade": "Capitaine", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "superviseur", "numero_employe": "CAP002", "date_embauche": "15/08/2020", "formations": [], "mot_de_passe": "TempPass123!"},
            
            # LIEUTENANTS
            {"nom": "Bernard", "prenom": "Pierre", "email": "employe@firemanager.ca", "telephone": "418-555-9999", "contact_urgence": "418-777-3333", "grade": "Lieutenant", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "LT001", "date_embauche": "21/09/2019", "formations": [], "mot_de_passe": "employe123"},
            {"nom": "Gagnon", "prenom": "Julie", "email": "julie.gagnon@firemanager.ca", "telephone": "514-666-7788", "contact_urgence": "514-999-3333", "grade": "Lieutenant", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "LT002", "date_embauche": "10/03/2021", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Roy", "prenom": "Alexandre", "email": "alex.roy@firemanager.ca", "telephone": "450-111-2222", "contact_urgence": "450-888-4444", "grade": "Lieutenant", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "LT003", "date_embauche": "05/11/2022", "formations": [], "mot_de_passe": "TempPass123!"},
            
            # POMPIERS TEMPS PLEIN
            {"nom": "Lavoie", "prenom": "Marc", "email": "marc.lavoie@firemanager.ca", "telephone": "514-777-8899", "contact_urgence": "514-666-5555", "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "POM001", "date_embauche": "12/04/2021", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Côté", "prenom": "David", "email": "david.cote@firemanager.ca", "telephone": "418-888-9900", "contact_urgence": "418-555-6666", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "POM002", "date_embauche": "28/09/2022", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Bouchard", "prenom": "Simon", "email": "simon.bouchard@firemanager.ca", "telephone": "514-999-1234", "contact_urgence": "514-777-7777", "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "POM003", "date_embauche": "16/01/2023", "formations": [], "mot_de_passe": "TempPass123!"},
            
            # POMPIERS TEMPS PARTIEL POUR DÉMO ASSIGNATION AUTO (12 employés)
            {"nom": "Garcia", "prenom": "Claire", "email": "partiel@firemanager.ca", "telephone": "514-888-9900", "contact_urgence": "514-666-4444", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 25, "role": "employe", "numero_employe": "PTP001", "date_embauche": "02/11/2023", "formations": [], "mot_de_passe": "partiel123"},
            {"nom": "Martin", "prenom": "Sarah", "email": "sarah.martin@firemanager.ca", "telephone": "450-555-6666", "contact_urgence": "450-999-8888", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 30, "role": "employe", "numero_employe": "PTP002", "date_embauche": "15/06/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Pelletier", "prenom": "Émilie", "email": "emilie.pelletier@firemanager.ca", "telephone": "418-333-7777", "contact_urgence": "418-666-9999", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 20, "role": "employe", "numero_employe": "PTP003", "date_embauche": "08/02/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Bergeron", "prenom": "Thomas", "email": "thomas.bergeron@firemanager.ca", "telephone": "514-444-8888", "contact_urgence": "514-333-9999", "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_partiel", "heures_max_semaine": 28, "role": "employe", "numero_employe": "PTP004", "date_embauche": "22/08/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Rousseau", "prenom": "Jessica", "email": "jessica.rousseau@firemanager.ca", "telephone": "514-777-1111", "contact_urgence": "514-888-2222", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 24, "role": "employe", "numero_employe": "PTP005", "date_embauche": "12/03/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Fournier", "prenom": "Antoine", "email": "antoine.fournier@firemanager.ca", "telephone": "418-555-2222", "contact_urgence": "418-777-3333", "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_partiel", "heures_max_semaine": 32, "role": "employe", "numero_employe": "PTP006", "date_embauche": "05/01/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Leclerc", "prenom": "Mathieu", "email": "mathieu.leclerc@firemanager.ca", "telephone": "450-666-3333", "contact_urgence": "450-888-4444", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 26, "role": "employe", "numero_employe": "PTP007", "date_embauche": "18/07/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Gauthier", "prenom": "Isabelle", "email": "isabelle.gauthier@firemanager.ca", "telephone": "514-999-4444", "contact_urgence": "514-666-5555", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 22, "role": "employe", "numero_employe": "PTP008", "date_embauche": "30/04/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Beaulieu", "prenom": "Nicolas", "email": "nicolas.beaulieu@firemanager.ca", "telephone": "418-444-5555", "contact_urgence": "418-777-6666", "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_partiel", "heures_max_semaine": 35, "role": "employe", "numero_employe": "PTP009", "date_embauche": "14/09/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Caron", "prenom": "Melissa", "email": "melissa.caron@firemanager.ca", "telephone": "514-333-6666", "contact_urgence": "514-999-7777", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 29, "role": "employe", "numero_employe": "PTP010", "date_embauche": "25/05/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Simard", "prenom": "Gabriel", "email": "gabriel.simard@firemanager.ca", "telephone": "450-777-7777", "contact_urgence": "450-333-8888", "grade": "Pompier", "fonction_superieur": True, "type_emploi": "temps_partiel", "heures_max_semaine": 27, "role": "employe", "numero_employe": "PTP011", "date_embauche": "03/11/2024", "formations": [], "mot_de_passe": "TempPass123!"},
            {"nom": "Mercier", "prenom": "Valérie", "email": "valerie.mercier@firemanager.ca", "telephone": "418-888-8888", "contact_urgence": "418-555-9999", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_partiel", "heures_max_semaine": 31, "role": "employe", "numero_employe": "PTP012", "date_embauche": "17/12/2023", "formations": [], "mot_de_passe": "TempPass123!"},
            
            # NOUVELLES RECRUES
            {"nom": "Morin", "prenom": "Kevin", "email": "kevin.morin@firemanager.ca", "telephone": "514-111-9999", "contact_urgence": "514-222-8888", "grade": "Pompier", "fonction_superieur": False, "type_emploi": "temps_plein", "heures_max_semaine": 40, "role": "employe", "numero_employe": "POM004", "date_embauche": "01/09/2024", "formations": [], "mot_de_passe": "TempPass123!"}
        ]
        
        # 2. FORMATIONS COMPLÈTES POUR CASERNE
        demo_formations = [
            {"nom": "Pompier 1", "description": "Formation de base pompier niveau 1 - Obligatoire pour tous", "duree_heures": 200, "validite_mois": 24, "obligatoire": True},
            {"nom": "Premiers Répondants", "description": "Formation premiers secours et réanimation", "duree_heures": 16, "validite_mois": 12, "obligatoire": True},
            {"nom": "Désincarcération", "description": "Techniques de désincarcération et sauvetage routier", "duree_heures": 24, "validite_mois": 36, "obligatoire": True},
            {"nom": "Classe 4A", "description": "Permis de conduire véhicules lourds et échelles", "duree_heures": 40, "validite_mois": 60, "obligatoire": False},
            {"nom": "Officier 2", "description": "Formation commandement et leadership", "duree_heures": 120, "validite_mois": 36, "obligatoire": False},
            {"nom": "Sauvetage Aquatique", "description": "Techniques de sauvetage en milieu aquatique", "duree_heures": 32, "validite_mois": 24, "obligatoire": False},
            {"nom": "Matières Dangereuses", "description": "Intervention matières dangereuses HAZMAT", "duree_heures": 48, "validite_mois": 36, "obligatoire": False},
            {"nom": "Sauvetage Technique", "description": "Sauvetage en espace clos et hauteur", "duree_heures": 56, "validite_mois": 24, "obligatoire": False}
        ]
        
        formation_ids = {}
        for formation_data in demo_formations:
            formation_obj = Formation(**formation_data)
            await db.formations.insert_one(formation_obj.dict())
            formation_ids[formation_data["nom"]] = formation_obj.id
        
        # 3. ASSIGNER FORMATIONS RÉALISTES PAR GRADE
        # Directeur - Toutes formations + Officier
        demo_users[0]["formations"] = [formation_ids["Pompier 1"], formation_ids["Officier 2"], formation_ids["Premiers Répondants"], formation_ids["Classe 4A"]]
        demo_users[1]["formations"] = [formation_ids["Pompier 1"], formation_ids["Officier 2"], formation_ids["Premiers Répondants"], formation_ids["Sauvetage Aquatique"]]
        
        # Capitaines - Formations supervision
        demo_users[2]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Désincarcération"], formation_ids["Classe 4A"]]
        demo_users[3]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Matières Dangereuses"]]
        
        # Lieutenants - Formations techniques
        demo_users[4]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Désincarcération"]]
        demo_users[5]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Sauvetage Aquatique"]]
        demo_users[6]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Sauvetage Technique"]]
        
        # Pompiers temps plein - Formations de base + spécialisations
        demo_users[7]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Désincarcération"]]
        demo_users[8]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"]]
        demo_users[9]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"], formation_ids["Classe 4A"]]
        
        # Pompiers temps partiel - Formations variables
        demo_users[10]["formations"] = [formation_ids["Pompier 1"]]
        demo_users[11]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"]]
        demo_users[12]["formations"] = [formation_ids["Pompier 1"]]
        demo_users[13]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"]]
        
        # Nouvelles recrues - Formations de base seulement
        demo_users[14]["formations"] = [formation_ids["Pompier 1"]]
        
        # Créer utilisateurs
        user_ids = {}
        for user_data in demo_users:
            user_dict = user_data.copy()
            user_dict["mot_de_passe_hash"] = get_password_hash(user_dict.pop("mot_de_passe"))
            user_dict["statut"] = "Actif"
            user_obj = User(**user_dict)
            await db.users.insert_one(user_obj.dict())
            user_ids[user_data["email"]] = user_obj.id
        
        # 4. CRÉER TYPES DE GARDE RÉALISTES
        demo_types_garde = [
            {"nom": "Garde Interne AM - Semaine", "heure_debut": "06:00", "heure_fin": "18:00", "personnel_requis": 4, "duree_heures": 12, "couleur": "#10B981", "jours_application": ["monday", "tuesday", "wednesday", "thursday", "friday"], "officier_obligatoire": True},
            {"nom": "Garde Interne PM - Semaine", "heure_debut": "18:00", "heure_fin": "06:00", "personnel_requis": 3, "duree_heures": 12, "couleur": "#3B82F6", "jours_application": ["monday", "tuesday", "wednesday", "thursday", "friday"], "officier_obligatoire": True},
            {"nom": "Garde Weekend Jour", "heure_debut": "08:00", "heure_fin": "20:00", "personnel_requis": 3, "duree_heures": 12, "couleur": "#F59E0B", "jours_application": ["saturday", "sunday"], "officier_obligatoire": True},
            {"nom": "Garde Weekend Nuit", "heure_debut": "20:00", "heure_fin": "08:00", "personnel_requis": 2, "duree_heures": 12, "couleur": "#8B5CF6", "jours_application": ["saturday", "sunday"], "officier_obligatoire": False},
            {"nom": "Garde Externe Citerne", "heure_debut": "00:00", "heure_fin": "23:59", "personnel_requis": 1, "duree_heures": 24, "couleur": "#EF4444", "jours_application": ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"], "officier_obligatoire": False}
        ]
        
        type_garde_ids = {}
        for type_garde_data in demo_types_garde:
            type_garde_obj = TypeGarde(**type_garde_data)
            await db.types_garde.insert_one(type_garde_obj.dict())
            type_garde_ids[type_garde_data["nom"]] = type_garde_obj.id
        
        # 5. CRÉER HISTORIQUE ASSIGNATIONS (6 semaines)
        assignations_created = 0
        users_list = list(user_ids.values())
        
        for week_offset in range(-6, 1):  # 6 semaines passées + courante
            week_start = datetime.now(timezone.utc).date() + timedelta(weeks=week_offset)
            week_start = week_start - timedelta(days=week_start.weekday())
            
            for day_offset in range(7):
                date_assignation = week_start + timedelta(days=day_offset)
                date_str = date_assignation.strftime("%Y-%m-%d")
                day_name = date_assignation.strftime("%A").lower()
                
                # Assigner gardes selon jours d'application
                for type_nom, type_id in type_garde_ids.items():
                    type_garde = next(t for t in demo_types_garde if t["nom"] == type_nom)
                    
                    if day_name in type_garde["jours_application"]:
                        # Assigner aléatoirement 70% des gardes
                        if assignations_created % 3 != 2:  # 2/3 des gardes assignées
                            import random
                            user_id = random.choice(users_list)
                            
                            assignation_obj = Assignation(
                                user_id=user_id,
                                type_garde_id=type_id,
                                date=date_str,
                                assignation_type="auto"
                            )
                            await db.assignations.insert_one(assignation_obj.dict())
                            assignations_created += 1
        
        # 6. CRÉER DISPONIBILITÉS MASSIVES POUR TEMPS PARTIEL (pour démo assignation auto)
        temps_partiel_users = [
            user_ids["partiel@firemanager.ca"],
            user_ids["sarah.martin@firemanager.ca"],
            user_ids["emilie.pelletier@firemanager.ca"],
            user_ids["thomas.bergeron@firemanager.ca"],
            user_ids["jessica.rousseau@firemanager.ca"],
            user_ids["antoine.fournier@firemanager.ca"],
            user_ids["mathieu.leclerc@firemanager.ca"],
            user_ids["isabelle.gauthier@firemanager.ca"],
            user_ids["nicolas.beaulieu@firemanager.ca"],
            user_ids["melissa.caron@firemanager.ca"],
            user_ids["gabriel.simard@firemanager.ca"],
            user_ids["valerie.mercier@firemanager.ca"]
        ]
        
        disponibilites_created = 0
        
        # Pour chaque employé temps partiel, créer des disponibilités variées
        for i, user_id in enumerate(temps_partiel_users):
            # Patterns de disponibilité différents pour variété
            if i % 4 == 0:  # Pattern 1: Lun-Mer-Ven
                jours_pattern = [0, 2, 4]  # Lundi, Mercredi, Vendredi
            elif i % 4 == 1:  # Pattern 2: Mar-Jeu-Sam
                jours_pattern = [1, 3, 5]  # Mardi, Jeudi, Samedi
            elif i % 4 == 2:  # Pattern 3: Mer-Ven-Dim
                jours_pattern = [2, 4, 6]  # Mercredi, Vendredi, Dimanche
            else:  # Pattern 4: Lun-Jeu-Sam
                jours_pattern = [0, 3, 5]  # Lundi, Jeudi, Samedi
            
            # Créer disponibilités pour 8 semaines (2 mois futurs)
            for week_offset in range(0, 8):
                week_start = datetime.now(timezone.utc).date() + timedelta(weeks=week_offset)
                week_start = week_start - timedelta(days=week_start.weekday())
                
                for day_offset in jours_pattern:
                    date_dispo = week_start + timedelta(days=day_offset)
                    
                    # Créer disponibilités pour différents types de garde
                    # 80% pour Garde Interne (semaine)
                    if day_offset < 5:  # Lundi-Vendredi
                        # Disponible pour garde AM ou PM (alternativement)
                        type_garde_am = type_garde_ids["Garde Interne AM - Semaine"]
                        type_garde_pm = type_garde_ids["Garde Interne PM - Semaine"]
                        
                        # Disponible pour garde AM
                        if disponibilites_created % 3 != 2:  # 66% de chance
                            dispo_obj = Disponibilite(
                                user_id=user_id,
                                date=date_dispo.strftime("%Y-%m-%d"),
                                type_garde_id=type_garde_am,
                                heure_debut="06:00",
                                heure_fin="18:00",
                                statut="disponible"
                            )
                            await db.disponibilites.insert_one(dispo_obj.dict())
                            disponibilites_created += 1
                        
                        # Disponible pour garde PM
                        if disponibilites_created % 4 != 3:  # 75% de chance
                            dispo_obj = Disponibilite(
                                user_id=user_id,
                                date=date_dispo.strftime("%Y-%m-%d"),
                                type_garde_id=type_garde_pm,
                                heure_debut="18:00",
                                heure_fin="06:00",
                                statut="disponible"
                            )
                            await db.disponibilites.insert_one(dispo_obj.dict())
                            disponibilites_created += 1
                    else:  # Weekend
                        # Disponible pour garde weekend
                        if disponibilites_created % 2 == 0:  # 50% de chance
                            type_garde_weekend = type_garde_ids["Garde Weekend Jour"]
                            dispo_obj = Disponibilite(
                                user_id=user_id,
                                date=date_dispo.strftime("%Y-%m-%d"),
                                type_garde_id=type_garde_weekend,
                                heure_debut="08:00",
                                heure_fin="20:00",
                                statut="disponible"
                            )
                            await db.disponibilites.insert_one(dispo_obj.dict())
                            disponibilites_created += 1
        
        # 7. CRÉER SESSIONS DE FORMATION
        demo_sessions = [
            {"titre": "Formation Sauvetage Aquatique - Niveau 1", "competence_id": formation_ids["Sauvetage Aquatique"], "duree_heures": 32, "date_debut": "2025-01-15", "heure_debut": "09:00", "lieu": "Piscine municipale", "formateur": "Capitaine Sarah Tremblay", "descriptif": "Formation complète aux techniques de sauvetage aquatique", "plan_cours": "", "places_max": 12, "participants": [], "statut": "planifie"},
            {"titre": "Perfectionnement Désincarcération", "competence_id": formation_ids["Désincarcération"], "duree_heures": 16, "date_debut": "2025-01-22", "heure_debut": "13:00", "lieu": "Centre formation sécurité", "formateur": "Lieutenant Pierre Bernard", "descriptif": "Perfectionnement techniques de désincarcération moderne", "plan_cours": "", "places_max": 8, "participants": [], "statut": "planifie"},
            {"titre": "Matières Dangereuses HAZMAT", "competence_id": formation_ids["Matières Dangereuses"], "duree_heures": 48, "date_debut": "2025-02-05", "heure_debut": "08:00", "lieu": "Centre HAZMAT Montréal", "formateur": "Expert externe - Dr. Martin Dubois", "descriptif": "Formation complète intervention matières dangereuses", "plan_cours": "", "places_max": 15, "participants": [], "statut": "planifie"}
        ]
        
        for session_data in demo_sessions:
            session_obj = SessionFormation(**session_data)
            await db.sessions_formation.insert_one(session_obj.dict())
        
        return {
            "message": "Données démo CLIENT créées avec succès",
            "details": {
                "utilisateurs": len(demo_users),
                "formations": len(demo_formations),
                "types_garde": len(demo_types_garde),
                "assignations_historiques": assignations_created,
                "disponibilites": disponibilites_created,
                "sessions_formation": len(demo_sessions),
                "breakdown": {
                    "admins": 1,
                    "superviseurs": 3,
                    "employes_temps_plein": 7,
                    "employes_temps_partiel": 12,
                    "total_personnel": 23
                }
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erreur: {str(e)}")

# Initialize demo data
@api_router.post("/init-demo-data")
async def init_demo_data():
    # Clear existing data
    await db.users.delete_many({})
    await db.types_garde.delete_many({})
    await db.assignations.delete_many({})
    await db.planning.delete_many({})
    await db.demandes_remplacement.delete_many({})
    
    # Create demo users
    demo_users = [
        {
            "nom": "Dupont",
            "prenom": "Jean",
            "email": "admin@firemanager.ca",
            "telephone": "514-111-2233",
            "contact_urgence": "514-999-1111",
            "grade": "Directeur",
            "type_emploi": "temps_plein",
            "heures_max_semaine": 40,  # Temps plein standard
            "role": "admin",
            "numero_employe": "ADM001",
            "date_embauche": "14/01/2020",
            "formations": [],
            "mot_de_passe": "admin123"
        },
        {
            "nom": "Dubois",
            "prenom": "Sophie",
            "email": "superviseur@firemanager.ca",
            "telephone": "514-444-5566",
            "contact_urgence": "514-888-2222",
            "grade": "Directeur",
            "type_emploi": "temps_plein",
            "heures_max_semaine": 40,  # Temps plein standard
            "role": "superviseur",
            "numero_employe": "POM001",
            "date_embauche": "07/01/2022",
            "formations": [],
            "mot_de_passe": "superviseur123"
        },
        {
            "nom": "Bernard",
            "prenom": "Pierre",
            "email": "employe@firemanager.ca",
            "telephone": "418-555-9999",
            "contact_urgence": "418-777-3333",
            "grade": "Capitaine",
            "type_emploi": "temps_plein",
            "heures_max_semaine": 40,  # Temps plein standard
            "role": "employe",
            "numero_employe": "POM002",
            "date_embauche": "21/09/2019",
            "formations": [],
            "mot_de_passe": "employe123"
        },
        {
            "nom": "Garcia",
            "prenom": "Claire",
            "email": "partiel@firemanager.ca",
            "telephone": "514-888-9900",
            "contact_urgence": "514-666-4444",
            "grade": "Pompier",
            "type_emploi": "temps_partiel",
            "heures_max_semaine": 25,  # 25h max par semaine
            "role": "employe",
            "numero_employe": "POM005",
            "date_embauche": "02/11/2020",
            "formations": [],
            "mot_de_passe": "partiel123"
        }
    ]
    
    # First create formations
    demo_formations = [
        {
            "nom": "Classe 4A",
            "description": "Formation de conduite véhicules lourds",
            "duree_heures": 40,
            "validite_mois": 60,
            "obligatoire": False
        },
        {
            "nom": "Désincarcération",
            "description": "Techniques de désincarcération",
            "duree_heures": 24,
            "validite_mois": 36,
            "obligatoire": True
        },
        {
            "nom": "Pompier 1",
            "description": "Formation de base pompier niveau 1",
            "duree_heures": 200,
            "validite_mois": 24,
            "obligatoire": True
        },
        {
            "nom": "Officier 2",
            "description": "Formation officier niveau 2",
            "duree_heures": 120,
            "validite_mois": 36,
            "obligatoire": False
        },
        {
            "nom": "Premiers Répondants",
            "description": "Formation premiers secours",
            "duree_heures": 16,
            "validite_mois": 12,
            "obligatoire": True
        }
    ]
    
    formation_ids = {}
    for formation_data in demo_formations:
        formation_obj = Formation(**formation_data)
        await db.formations.insert_one(formation_obj.dict())
        formation_ids[formation_data["nom"]] = formation_obj.id
    
    # Update users with formation IDs
    demo_users[0]["formations"] = [formation_ids["Officier 2"], formation_ids["Pompier 1"]]  # Jean
    demo_users[1]["formations"] = [formation_ids["Pompier 1"], formation_ids["Premiers Répondants"]]  # Sophie  
    demo_users[2]["formations"] = [formation_ids["Classe 4A"], formation_ids["Désincarcération"]]  # Pierre
    demo_users[3]["formations"] = []  # Claire - aucune formation
    
    for user_data in demo_users:
        user_dict = user_data.copy()
        user_dict["mot_de_passe_hash"] = get_password_hash(user_dict.pop("mot_de_passe"))
        user_dict["statut"] = "Actif"
        user_obj = User(**user_dict)
        await db.users.insert_one(user_obj.dict())
    
    # Create demo garde types
    demo_types_garde = [
        {
            "nom": "Garde Interne AM - Semaine",
            "heure_debut": "06:00",
            "heure_fin": "12:00",
            "personnel_requis": 4,
            "duree_heures": 6,
            "couleur": "#10B981",
            "jours_application": ["monday", "tuesday", "wednesday", "thursday", "friday"],
            "officier_obligatoire": True
        },
        {
            "nom": "Garde Interne PM - Semaine",
            "heure_debut": "12:00",
            "heure_fin": "18:00",
            "personnel_requis": 4,
            "duree_heures": 6,
            "couleur": "#3B82F6",
            "jours_application": ["monday", "tuesday", "wednesday", "thursday", "friday"],
            "officier_obligatoire": True
        },
        {
            "nom": "Garde Externe Citerne",
            "heure_debut": "18:00",
            "heure_fin": "06:00",
            "personnel_requis": 1,
            "duree_heures": 12,
            "couleur": "#8B5CF6",
            "jours_application": ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"],
            "officier_obligatoire": False
        }
    ]
    
    for type_garde_data in demo_types_garde:
        type_garde_obj = TypeGarde(**type_garde_data)
        await db.types_garde.insert_one(type_garde_obj.dict())
    
    # Create demo disponibilités for part-time employee (Claire Garcia) with specific dates
    claire_user = await db.users.find_one({"email": "partiel@firemanager.ca"})
    if claire_user:
        # Create availabilities for next 2 weeks
        today = datetime.now(timezone.utc).date()
        demo_disponibilites = []
        
        # Generate availabilities for specific dates
        for week_offset in range(4):  # Next 4 weeks
            week_start = today + timedelta(weeks=week_offset)
            week_start = week_start - timedelta(days=week_start.weekday())  # Get Monday
            
            # Claire is available Monday, Wednesday, Friday
            for day_offset in [0, 2, 4]:  # Monday, Wednesday, Friday
                date_available = week_start + timedelta(days=day_offset)
                demo_disponibilites.append({
                    "user_id": claire_user["id"],
                    "date": date_available.strftime("%Y-%m-%d"),
                    "heure_debut": "08:00",
                    "heure_fin": "16:00",
                    "statut": "disponible"
                })
        
        for dispo_data in demo_disponibilites:
            dispo_obj = Disponibilite(**dispo_data)
            await db.disponibilites.insert_one(dispo_obj.dict())
    
    return {"message": "Données de démonstration créées avec succès"}

# ==================== NOTIFICATIONS ====================

@api_router.get("/notifications", response_model=List[Notification])
async def get_notifications(current_user: User = Depends(get_current_user)):
    """Récupère toutes les notifications de l'utilisateur connecté"""
    notifications = await db.notifications.find(
        {"destinataire_id": current_user.id}
    ).sort("date_creation", -1).limit(50).to_list(50)
    
    cleaned_notifications = [clean_mongo_doc(notif) for notif in notifications]
    return [Notification(**notif) for notif in cleaned_notifications]

@api_router.get("/notifications/non-lues/count")
async def get_unread_count(current_user: User = Depends(get_current_user)):
    """Compte le nombre de notifications non lues"""
    count = await db.notifications.count_documents({
        "destinataire_id": current_user.id,
        "statut": "non_lu"
    })
    return {"count": count}

@api_router.put("/notifications/{notification_id}/marquer-lu")
async def marquer_notification_lue(notification_id: str, current_user: User = Depends(get_current_user)):
    """Marque une notification comme lue"""
    notification = await db.notifications.find_one({
        "id": notification_id,
        "destinataire_id": current_user.id
    })
    
    if not notification:
        raise HTTPException(status_code=404, detail="Notification non trouvée")
    
    await db.notifications.update_one(
        {"id": notification_id},
        {"$set": {
            "statut": "lu",
            "date_lecture": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": "Notification marquée comme lue"}

@api_router.put("/notifications/marquer-toutes-lues")
async def marquer_toutes_lues(current_user: User = Depends(get_current_user)):
    """Marque toutes les notifications comme lues"""
    result = await db.notifications.update_many(
        {
            "destinataire_id": current_user.id,
            "statut": "non_lu"
        },
        {"$set": {
            "statut": "lu",
            "date_lecture": datetime.now(timezone.utc).isoformat()
        }}
    )
    
    return {"message": f"{result.modified_count} notification(s) marquée(s) comme lue(s)"}

# Helper function pour créer des notifications
async def creer_notification(
    destinataire_id: str,
    type: str,
    titre: str,
    message: str,
    lien: Optional[str] = None,
    data: Optional[Dict[str, Any]] = None
):
    """Crée une notification dans la base de données"""
    notification = Notification(
        destinataire_id=destinataire_id,
        type=type,
        titre=titre,
        message=message,
        lien=lien,
        data=data or {}
    )
    await db.notifications.insert_one(notification.dict())
    return notification

# ==================== PARAMÈTRES REMPLACEMENTS ====================

@api_router.get("/parametres/remplacements")
async def get_parametres_remplacements(current_user: User = Depends(get_current_user)):
    """Récupère les paramètres de remplacements"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    parametres = await db.parametres_remplacements.find_one()
    
    if not parametres:
        # Créer paramètres par défaut
        default_params = ParametresRemplacements()
        await db.parametres_remplacements.insert_one(default_params.dict())
        return default_params
    
    cleaned_params = clean_mongo_doc(parametres)
    return ParametresRemplacements(**cleaned_params)

@api_router.put("/parametres/remplacements")
async def update_parametres_remplacements(
    parametres: ParametresRemplacements,
    current_user: User = Depends(get_current_user)
):
    """Met à jour les paramètres de remplacements"""
    if current_user.role not in ["admin"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    existing = await db.parametres_remplacements.find_one()
    
    if existing:
        await db.parametres_remplacements.update_one(
            {"id": existing["id"]},
            {"$set": parametres.dict()}
        )
    else:
        await db.parametres_remplacements.insert_one(parametres.dict())
    
    return {"message": "Paramètres mis à jour avec succès"}

# ==================== EPI ROUTES ====================

@api_router.post("/epi", response_model=EPIEmploye)
async def create_epi(epi: EPIEmployeCreate, current_user: User = Depends(get_current_user)):
    """Crée un nouvel EPI pour un employé (Admin/Superviseur)"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Vérifier si l'employé existe
    employe = await db.users.find_one({"id": epi.employe_id})
    if not employe:
        raise HTTPException(status_code=404, detail="Employé non trouvé")
    
    # Vérifier qu'il n'existe pas déjà un EPI de ce type pour cet employé
    existing_epi = await db.epi_employes.find_one({
        "employe_id": epi.employe_id,
        "type_epi": epi.type_epi
    })
    
    if existing_epi:
        raise HTTPException(
            status_code=400,
            detail=f"Un EPI de type {epi.type_epi} existe déjà pour cet employé"
        )
    
    epi_obj = EPIEmploye(**epi.dict())
    await db.epi_employes.insert_one(epi_obj.dict())
    
    return epi_obj

@api_router.get("/epi/employe/{employe_id}", response_model=List[EPIEmploye])
async def get_epi_employe(employe_id: str, current_user: User = Depends(get_current_user)):
    """Récupère tous les EPI d'un employé"""
    # Admin et superviseur peuvent voir tous les EPI
    # Employé peut voir seulement ses propres EPI
    if current_user.role not in ["admin", "superviseur"] and current_user.id != employe_id:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    epis = await db.epi_employes.find({"employe_id": employe_id}).to_list(100)
    cleaned_epis = [clean_mongo_doc(epi) for epi in epis]
    return [EPIEmploye(**epi) for epi in cleaned_epis]

@api_router.get("/epi/{epi_id}", response_model=EPIEmploye)
async def get_epi_by_id(epi_id: str, current_user: User = Depends(get_current_user)):
    """Récupère un EPI spécifique par son ID"""
    epi = await db.epi_employes.find_one({"id": epi_id})
    
    if not epi:
        raise HTTPException(status_code=404, detail="EPI non trouvé")
    
    # Vérifier les permissions
    if current_user.role not in ["admin", "superviseur"] and current_user.id != epi["employe_id"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    cleaned_epi = clean_mongo_doc(epi)
    return EPIEmploye(**cleaned_epi)

@api_router.put("/epi/{epi_id}", response_model=EPIEmploye)
async def update_epi(
    epi_id: str,
    epi_update: EPIEmployeUpdate,
    current_user: User = Depends(get_current_user)
):
    """
    Met à jour un EPI
    - Admin/Superviseur: peuvent modifier tous les champs
    - Employé: peut modifier uniquement la taille
    """
    epi = await db.epi_employes.find_one({"id": epi_id})
    
    if not epi:
        raise HTTPException(status_code=404, detail="EPI non trouvé")
    
    # Vérifier les permissions
    is_owner = current_user.id == epi["employe_id"]
    is_admin_or_supervisor = current_user.role in ["admin", "superviseur"]
    
    if not (is_owner or is_admin_or_supervisor):
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Si c'est un employé, il ne peut modifier que la taille
    if not is_admin_or_supervisor:
        if epi_update.etat or epi_update.date_expiration or epi_update.date_prochaine_inspection or epi_update.notes:
            raise HTTPException(
                status_code=403,
                detail="Les employés peuvent modifier uniquement la taille"
            )
    
    # Préparer les champs à mettre à jour
    update_data = {k: v for k, v in epi_update.dict().items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    await db.epi_employes.update_one(
        {"id": epi_id},
        {"$set": update_data}
    )
    
    updated_epi = await db.epi_employes.find_one({"id": epi_id})
    cleaned_epi = clean_mongo_doc(updated_epi)
    return EPIEmploye(**cleaned_epi)

@api_router.delete("/epi/{epi_id}")
async def delete_epi(epi_id: str, current_user: User = Depends(get_current_user)):
    """Supprime un EPI (Admin/Superviseur seulement)"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    result = await db.epi_employes.delete_one({"id": epi_id})
    
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="EPI non trouvé")
    
    return {"message": "EPI supprimé avec succès"}

@api_router.post("/epi/{epi_id}/inspection")
async def add_inspection(
    epi_id: str,
    inspection: InspectionEPI,
    current_user: User = Depends(get_current_user)
):
    """Ajoute une inspection à un EPI"""
    epi = await db.epi_employes.find_one({"id": epi_id})
    
    if not epi:
        raise HTTPException(status_code=404, detail="EPI non trouvé")
    
    # L'inspection peut être faite par l'employé lui-même ou par un admin/superviseur
    if current_user.id != epi["employe_id"] and current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Ajouter l'inspection à l'historique
    inspection_data = inspection.dict()
    inspection_data["inspecteur_id"] = current_user.id
    
    await db.epi_employes.update_one(
        {"id": epi_id},
        {
            "$push": {"historique_inspections": inspection_data},
            "$set": {"updated_at": datetime.now(timezone.utc).isoformat()}
        }
    )
    
    # Si l'inspection indique un remplacement nécessaire, créer une notification
    if inspection.resultat == "Remplacement nécessaire":
        employe = await db.users.find_one({"id": epi["employe_id"]})
        
        # Notifier les admins et superviseurs
        superviseurs_admins = await db.users.find({"role": {"$in": ["superviseur", "admin"]}}).to_list(100)
        for superviseur in superviseurs_admins:
            await creer_notification(
                destinataire_id=superviseur["id"],
                type="epi_remplacement",
                titre="Remplacement EPI nécessaire",
                message=f"{employe['prenom']} {employe['nom']} signale que son {epi['type_epi']} nécessite un remplacement",
                lien="/personnel",
                data={"epi_id": epi_id, "employe_id": epi["employe_id"]}
            )
    
    return {"message": "Inspection ajoutée avec succès"}

@api_router.get("/epi/alertes/all", response_model=List[Dict[str, Any]])
async def get_epi_alerts(current_user: User = Depends(get_current_user)):
    """Récupère toutes les alertes EPI (expirations et inspections à venir)"""
    if current_user.role not in ["admin", "superviseur"]:
        raise HTTPException(status_code=403, detail="Accès refusé")
    
    # Récupérer tous les EPI
    all_epis = await db.epi_employes.find().to_list(1000)
    
    alerts = []
    today = datetime.now(timezone.utc)
    
    for epi in all_epis:
        # Récupérer l'info employé
        employe = await db.users.find_one({"id": epi["employe_id"]})
        if not employe:
            continue
        
        # Vérifier expiration
        if epi.get("date_expiration"):
            date_exp = datetime.fromisoformat(epi["date_expiration"].replace('Z', '+00:00'))
            days_until_expiration = (date_exp - today).days
            
            if days_until_expiration <= 30:
                alerts.append({
                    "type": "expiration",
                    "epi_id": epi["id"],
                    "employe_id": epi["employe_id"],
                    "employe_nom": f"{employe['prenom']} {employe['nom']}",
                    "type_epi": epi["type_epi"],
                    "jours_restants": days_until_expiration,
                    "date_expiration": epi["date_expiration"],
                    "priorite": "haute" if days_until_expiration <= 7 else "moyenne"
                })
        
        # Vérifier inspection
        if epi.get("date_prochaine_inspection"):
            date_insp = datetime.fromisoformat(epi["date_prochaine_inspection"].replace('Z', '+00:00'))
            days_until_inspection = (date_insp - today).days
            
            if days_until_inspection <= 14:
                alerts.append({
                    "type": "inspection",
                    "epi_id": epi["id"],
                    "employe_id": epi["employe_id"],
                    "employe_nom": f"{employe['prenom']} {employe['nom']}",
                    "type_epi": epi["type_epi"],
                    "jours_restants": days_until_inspection,
                    "date_inspection": epi["date_prochaine_inspection"],
                    "priorite": "haute" if days_until_inspection <= 3 else "moyenne"
                })
    
    return alerts

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()